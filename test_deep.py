"""
בדיקות עומק — מקרים שtest_all.py לא מכסה:
  - גבולות חישוב תאריך (רביעי ביום רביעי עצמו)
  - migration של DB ישן (עמודת weekly_status)
  - דיוק נתוני סיכום
  - מחיקה כוללת עם cascade
  - רוטציית גיבויים (מגבלת 30)
  - תאימות unicode / עברית עם ניקוד
  - ביצועי DB עם 200 רשומות
  - גיבוי WAL integrity (לאחר תיקון)
  - שמירת הגדרות בין חיבורים
  - ייצוא Excel — שם קובץ + תוכן
"""
import sys, os, tempfile, shutil, sqlite3, time
sys.stdout.reconfigure(encoding="utf-8")
os.environ["QT_QPA_PLATFORM"] = "offscreen"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# DB זמני
_tmp_fd, _TMP_DB = tempfile.mkstemp(suffix=".db")
os.close(_tmp_fd)
import database as db
db.DB_PATH = _TMP_DB
from pathlib import Path
assert Path(db.DB_PATH).name != "data.db"

_errors: list = []
_passed = 0

def check(name: str, condition: bool, detail: str = ""):
    global _passed
    if condition:
        print(f"  ✓ {name}")
        _passed += 1
    else:
        print(f"  ✗ {name} {detail}")
        _errors.append(name)


# ══════════════════════════════════════════════════
# רובד A — גבולות חישוב תאריך
# ══════════════════════════════════════════════════
print("\n=== A: גבולות חישוב תאריך ===")
from database import calculate_next_dist, next_wednesday
from datetime import date, timedelta

# כשהיום הוא רביעי — הרביעי הבא הוא +7 ימים (לא היום)
wed = date(2026, 6, 3)   # רביעי ידוע
assert wed.weekday() == 2
d = next_wednesday(wed)
check("next_wednesday(רביעי) → +7 (לא היום)", (d - wed).days == 7, f"got {d}")

# כשהיום הוא שלישי — הרביעי הבא הוא +1
tue = date(2026, 6, 2)
d2 = next_wednesday(tue)
check("next_wednesday(שלישי) → +1", (d2 - tue).days == 1, f"got {d2}")

# כשהיום הוא חמישי — הרביעי הבא הוא +6
thu = date(2026, 6, 4)
d3 = next_wednesday(thu)
check("next_wednesday(חמישי) → +6", (d3 - thu).days == 6, f"got {d3}")

# calculate_next_dist: שבועי מרביעי → רביעי הבא
d4 = calculate_next_dist("2026-06-03", "שבועי")   # מרביעי שבועי
check("שבועי מרביעי → +7", d4.weekday() == 2 and (d4 - wed).days == 7)

# calculate_next_dist: דו-שבועי
d5 = calculate_next_dist("2026-06-03", "דו-שבועי")
check("דו-שבועי ≥ +13 ימים", (d5 - wed).days >= 13)

# calculate_next_dist: חודשי
d6 = calculate_next_dist("2026-06-03", "חודשי")
check("חודשי ≥ +29 ימים", (d6 - wed).days >= 29)

# calculate_next_dist: חד-פעמי → רביעי הקרוב מהיום (לא מתאריך ניתן)
d7 = calculate_next_dist("2026-06-03", "חד-פעמי")
check("חד-פעמי → רביעי הקרוב מהיום", d7.weekday() == 2)

# ══════════════════════════════════════════════════
# רובד B — Migration: DB ישן בלי עמודת weekly_status
# ══════════════════════════════════════════════════
print("\n=== B: Migration DB ישן ===")
fd2, old_db = tempfile.mkstemp(suffix=".db"); os.close(fd2)
old_conn = sqlite3.connect(old_db)
old_conn.executescript("""
    CREATE TABLE recipients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        full_name TEXT NOT NULL,
        status TEXT DEFAULT 'פעיל',
        frequency TEXT DEFAULT '',
        last_distribution TEXT,
        next_distribution TEXT,
        phone1 TEXT, phone2 TEXT, phone3 TEXT,
        address TEXT, area TEXT DEFAULT '',
        souls INTEGER DEFAULT 0,
        start_date TEXT, notes TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE settings (key TEXT PRIMARY KEY, value TEXT);
    INSERT INTO settings VALUES ('password', '1234');
    INSERT INTO settings VALUES ('backup_folder', '');
    INSERT INTO settings VALUES ('last_backup_at', '');
    INSERT INTO recipients (full_name, status) VALUES ('ישן', 'פעיל');
""")
old_conn.close()

# הפעל init_db על DB ישן → צריך להוסיף weekly_status
orig_path = db.DB_PATH
db.DB_PATH = old_db
db.init_db()

_mc = sqlite3.connect(old_db)
cols = {r[1] for r in _mc.execute("PRAGMA table_info(recipients)")}  # col[1] = name
rows = _mc.execute("SELECT full_name FROM recipients").fetchall()
_mc.close()
check("weekly_status column added by migration", "weekly_status" in cols)
check("existing data preserved after migration", any(r[0] == "ישן" for r in rows))

db.DB_PATH = orig_path
try: os.unlink(old_db)
except PermissionError: pass   # WAL keeps file open on Windows


# ══════════════════════════════════════════════════
# רובד C — דיוק נתוני סיכום
# ══════════════════════════════════════════════════
print("\n=== C: דיוק נתוני סיכום ===")
db.init_db()

ids = []
for i in range(5):
    rid = db.add_recipient({"full_name": f"פעיל {i}", "status": "פעיל",
                            "frequency": "שבועי", "souls": i+1})
    ids.append(rid)

db.add_recipient({"full_name": "מושהה", "status": "מושהה", "souls": 10})
db.add_recipient({"full_name": "הסתיים", "status": "הסתיים", "souls": 10})

# הוסף חלוקה לשניים
db.bulk_add_distributions(
    [{"id": ids[0], "full_name": "פעיל 0", "frequency": "שבועי"},
     {"id": ids[1], "full_name": "פעיל 1", "frequency": "שבועי"}],
    "2026-06-01", "מזון", 1, ""
)

stats = db.get_summary()
check("active count = 5", stats["active"] == 5, f"got {stats['active']}")
check("suspended count = 1", stats["suspended"] == 1, f"got {stats['suspended']}")
check("ended count = 1", stats["ended"] == 1, f"got {stats['ended']}")
check("total_souls = 1+2+3+4+5 = 15", stats["total_souls"] == 15, f"got {stats['total_souls']}")
check("dists_total = 2", stats["dists_total"] == 2, f"got {stats['dists_total']}")


# ══════════════════════════════════════════════════
# רובד D — force_delete cascade מלא
# ══════════════════════════════════════════════════
print("\n=== D: force_delete cascade ===")
rid_del = db.add_recipient({"full_name": "למחיקה", "status": "פעיל", "frequency": "שבועי"})
db.bulk_add_distributions(
    [{"id": rid_del, "full_name": "למחיקה", "frequency": "שבועי"}],
    "2026-06-01", "עוף", 1, ""
)
db.update_recipient(rid_del, {"status": "מושהה"})  # יוצר change_log

conn_check = sqlite3.connect(db.DB_PATH)
dist_before = conn_check.execute(
    "SELECT COUNT(*) FROM distributions WHERE recipient_id=?", (rid_del,)).fetchone()[0]
log_before = conn_check.execute(
    "SELECT COUNT(*) FROM change_log WHERE recipient_id=?", (rid_del,)).fetchone()[0]
conn_check.close()

check("dist before force_delete > 0", dist_before > 0, f"({dist_before})")
check("change_log before force_delete > 0", log_before > 0, f"({log_before})")

db.force_delete_recipient(rid_del)

conn_check2 = sqlite3.connect(db.DB_PATH)
rec_after   = conn_check2.execute("SELECT id FROM recipients WHERE id=?", (rid_del,)).fetchone()
dist_after  = conn_check2.execute(
    "SELECT COUNT(*) FROM distributions WHERE recipient_id=?", (rid_del,)).fetchone()[0]
log_after   = conn_check2.execute(
    "SELECT COUNT(*) FROM change_log WHERE recipient_id=?", (rid_del,)).fetchone()[0]
conn_check2.close()

check("recipient deleted", rec_after is None)
check("distributions deleted", dist_after == 0, f"({dist_after} remain)")
check("change_log deleted", log_after == 0, f"({log_after} remain)")


# ══════════════════════════════════════════════════
# רובד E — רוטציית גיבויים (מגבלת 30)
# ══════════════════════════════════════════════════
print("\n=== E: רוטציית גיבויים ===")
bk_dir = tempfile.mkdtemp()
db.set_setting("backup_folder", bk_dir)

from utils.backup import auto_backup

# צור 32 גיבויים רצופים
for i in range(32):
    time.sleep(0.02)   # הפרש זמן כדי לוודא שמות שונים
    auto_backup()

bk_files = sorted([f for f in os.listdir(bk_dir) if f.startswith("backup_")])
check("backup rotation keeps ≤30 files", len(bk_files) <= 30, f"({len(bk_files)} files)")
check("newest backup exists", len(bk_files) > 0)
db.set_setting("backup_folder", "")
shutil.rmtree(bk_dir, ignore_errors=True)


# ══════════════════════════════════════════════════
# רובד F — WAL Backup Integrity (לאחר תיקון)
# ══════════════════════════════════════════════════
print("\n=== F: WAL Backup Integrity ===")
fd3, wal_db = tempfile.mkstemp(suffix=".db"); os.close(fd3)
orig_path2 = db.DB_PATH
db.DB_PATH = wal_db
db.init_db()

# הוסף נתונים (ישבו ב-WAL לפני checkpoint)
wal_rid = db.add_recipient({"full_name": "WAL test", "souls": 7, "status": "פעיל"})
db.set_setting("last_backup_at", "")  # נקה

bk_dir2 = tempfile.mkdtemp()
db.set_setting("backup_folder", bk_dir2)
result = auto_backup()
check("WAL backup returns True", result is True)

bk_files2 = [f for f in os.listdir(bk_dir2) if f.startswith("backup_")]
if bk_files2:
    bk_path = os.path.join(bk_dir2, bk_files2[0])
    bk_conn = sqlite3.connect(bk_path)
    rows_bk = bk_conn.execute(
        "SELECT full_name, souls FROM recipients WHERE full_name='WAL test'").fetchall()
    bk_conn.close()
    check("WAL data in backup", len(rows_bk) == 1, f"({len(rows_bk)} rows)")
    check("WAL souls preserved", rows_bk[0][1] == 7 if rows_bk else False)
else:
    check("backup file created", False)

db.DB_PATH = orig_path2
shutil.rmtree(bk_dir2, ignore_errors=True)
try: os.unlink(wal_db)
except: pass


# ══════════════════════════════════════════════════
# רובד G — Unicode / עברית עם ניקוד ותווים מיוחדים
# ══════════════════════════════════════════════════
print("\n=== G: Unicode / עברית מיוחדת ===")
unicode_cases = [
    ("שָׁם עִם נִקּוּד", "050-1234567"),           # ניקוד
    ("Ó'Brien מקבל", "050-9876543"),              # Latin mixed
    ("מקבל עם 'גרשיים'", "052-1111111"),           # apostrophe
    ("אברהם    רבה", "053-2222222"),               # spaces
    ("🕍 ישראל", "054-3333333"),                   # emoji
]
for name, phone in unicode_cases:
    try:
        uid = db.add_recipient({"full_name": name, "phone1": phone, "status": "פעיל"})
        rec = db.get_recipient(uid)
        check(f"unicode name stored: {name[:12]}...", rec and rec["full_name"] == name)
    except Exception as e:
        check(f"unicode name stored: {name[:12]}...", False, str(e))

# חיפוש ב-get_all_recipients עם שם unicode
all_recs = db.get_all_recipients()
nikud_found = any("נִקּוּד" in r["full_name"] for r in all_recs)
check("ניקוד searchable in get_all_recipients", nikud_found)


# ══════════════════════════════════════════════════
# רובד H — ביצועי DB עם 200 רשומות
# ══════════════════════════════════════════════════
print("\n=== H: ביצועים עם 200 רשומות ===")
rows_to_add = [
    {"full_name": f"מקבל {i:03d}", "status": "פעיל",
     "frequency": ["שבועי","דו-שבועי","חודשי"][i % 3],
     "area": ["בעלז","נתיב",""][i % 3],
     "souls": (i % 5) + 1}
    for i in range(200)
]
t0 = time.perf_counter()
added, updated, conflicts = db.import_recipients_from_list(rows_to_add)
t1 = time.perf_counter()
check("200 rows imported", added == 200, f"({added})")
check("import 200 rows < 1 sec", (t1 - t0) < 1.0, f"({(t1-t0)*1000:.0f}ms)")

t2 = time.perf_counter()
all_r = db.get_all_recipients()
t3 = time.perf_counter()
check("get_all_recipients < 50ms", (t3 - t2) < 0.05, f"({(t3-t2)*1000:.0f}ms)")

t4 = time.perf_counter()
weekly = db.get_weekly_list(365)
t5 = time.perf_counter()
check("get_weekly_list < 100ms", (t5 - t4) < 0.1, f"({(t5-t4)*1000:.0f}ms)")
check("weekly list not empty", len(weekly) > 0, f"({len(weekly)})")


# ══════════════════════════════════════════════════
# רובד I — הגדרות שורדות חיבור חדש
# ══════════════════════════════════════════════════
print("\n=== I: הגדרות שורדות חיבור חדש ===")
db.set_setting("test_persist", "ערך_לשמירה")
# פתח חיבור חדש לאותו DB
conn_new = sqlite3.connect(db.DB_PATH)
conn_new.row_factory = sqlite3.Row
val = conn_new.execute("SELECT value FROM settings WHERE key='test_persist'").fetchone()
conn_new.close()
check("setting persists in new connection", val and val["value"] == "ערך_לשמירה")

# הגדרת סיסמא ברירת מחדל קיימת (מאוחסנת כ-hash PBKDF2, מאומתת דרך verify_password)
check("default password = 1234 persists", db.verify_password("1234"))


# ══════════════════════════════════════════════════
# רובד J — Excel Export: שם קובץ ותוכן
# ══════════════════════════════════════════════════
print("\n=== J: Excel Export קובץ ותוכן ===")
from utils.excel_utils import export_distribution_to_excel
import openpyxl

test_recs = [
    {"full_name": "כהן ראובן",  "phone1": "050-111", "area": "בעלז", "souls": 3},
    {"full_name": "לוי שמעון",  "phone1": "050-222", "area": "נתיב", "souls": 5},
]

path = export_distribution_to_excel(test_recs, "04/06/2026")
p = Path(path)
check("export file exists", p.exists())
check("filename has no spaces", " " not in p.name)
check("file in exports/ folder", "exports" in str(p))

# בדוק תוכן הקובץ
if p.exists():
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # שורה 1 = כותרת, שורה 2 = header עמודות, שורות 3+ = נתונים
    names_in_file = [ws.cell(r, 2).value for r in range(3, ws.max_row + 1)]
    check("both names in Excel", "כהן ראובן" in names_in_file and "לוי שמעון" in names_in_file,
          f"found: {names_in_file}")
    p.unlink()


# ══════════════════════════════════════════════════
# רובד K — שמירת weekly_status ב-DB
# ══════════════════════════════════════════════════
print("\n=== K: שמירת weekly_status ===")
wrid = db.add_recipient({"full_name": "שבועי טסט", "status": "פעיל", "frequency": "שבועי"})
db.update_recipient(wrid, {"weekly_status": "✓"})
rec_w = db.get_recipient(wrid)
check("weekly_status saved", rec_w and rec_w["weekly_status"] == "✓",
      f"got {rec_w.get('weekly_status') if rec_w else 'None'}")

# אפס
db.update_recipient(wrid, {"weekly_status": ""})
rec_w2 = db.get_recipient(wrid)
check("weekly_status clearable", rec_w2 and rec_w2["weekly_status"] == "")

# בדוק שה-change_log לא רושם weekly_status (אינו שדה מעוקב)
log_before = len(db.get_change_log())
db.update_recipient(wrid, {"weekly_status": "✗"})
log_after = len(db.get_change_log())
check("weekly_status NOT logged in change_log", log_after == log_before)


# ══════════════════════════════════════════════════
# רובד L — get_one_time_list: מקבל ללא last_distribution
# ══════════════════════════════════════════════════
print("\n=== L: חד-פעמי ללא תאריך קודם ===")
ot1 = db.add_recipient({"full_name": "חד ללא תאריך", "frequency": "חד-פעמי",
                         "status": "פעיל", "souls": 3})
ot2 = db.add_recipient({"full_name": "חד עם תאריך",  "frequency": "חד-פעמי",
                         "status": "פעיל", "souls": 5, "last_distribution": "2025-01-01"})

ot_list = db.get_one_time_list()
ot_names = [r["full_name"] for r in ot_list]

check("both one-time appear", "חד ללא תאריך" in ot_names and "חד עם תאריך" in ot_names)

# ללא תאריך → date(2000,1,1) → הכי ישן → ראשון
idx_no_date   = next(i for i, r in enumerate(ot_list) if r["full_name"] == "חד ללא תאריך")
idx_with_date = next(i for i, r in enumerate(ot_list) if r["full_name"] == "חד עם תאריך")
check("no-date appears before 2025 date", idx_no_date < idx_with_date,
      f"idx_no={idx_no_date}, idx_with={idx_with_date}")


# ══════════════════════════════════════════════════
# רובד M — get_distributions filter by name
# ══════════════════════════════════════════════════
print("\n=== M: get_distributions לפי שם ===")
target_id = db.add_recipient({"full_name": "מטרה לסינון", "status": "פעיל", "frequency": "שבועי"})
other_id  = db.add_recipient({"full_name": "אחר לסינון",  "status": "פעיל", "frequency": "שבועי"})
db.bulk_add_distributions(
    [{"id": target_id, "full_name": "מטרה לסינון", "frequency": "שבועי"}],
    "2026-06-01", "מזון", 1, ""
)
db.bulk_add_distributions(
    [{"id": other_id, "full_name": "אחר לסינון", "frequency": "שבועי"}],
    "2026-06-01", "עוף", 2, ""
)

dists_target = db.get_distributions(recipient_name="מטרה לסינון")
dists_other  = db.get_distributions(recipient_name="אחר לסינון")

check("filter by name — target only", all(d["recipient_name"] == "מטרה לסינון" for d in dists_target))
check("filter by name — other only",  all(d["recipient_name"] == "אחר לסינון"  for d in dists_other))
check("no cross-contamination", len(dists_target) >= 1 and len(dists_other) >= 1)


# ══════════════════════════════════════════════════
# רובד N — משקלי ניקוד מתכווננים
# ══════════════════════════════════════════════════
print("\n=== N: משקלי ניקוד מתכווננים ===")
db.set_need_weights(db.DEFAULT_NEED_WEIGHTS)
_wd = db.get_need_weights()
check("default weights: money=34", _wd["money"] == 34.0, f"got {_wd['money']}")
check("default weights: new fields 0", _wd["income"] == 0 and _wd["children"] == 0)
check("money parser '5,000 ₪' → 5000", db._need_num("5,000 ₪", "money") == 5000.0)
db.set_need_weights({"income": -3})
check("negative weight clamped to 0", db.get_need_weights()["income"] == 0.0)

# the chosen weight drives the ranking
db.reset_all_data()
db.add_recipient({"full_name": "גדולה", "frequency": "חד-פעמי", "status": "פעיל",
                  "priority": 3, "souls": 12, "income": "9000"})
db.add_recipient({"full_name": "עניה", "frequency": "חד-פעמי", "status": "פעיל",
                  "priority": 3, "souls": 3, "income": "1000"})
db.set_need_weights({"souls": 100, "money": 0, "recency": 0,
                     "income": 0, "housing": 0, "medical": 0, "children": 0})
check("souls-weighted → big family first", db.get_one_time_list()[0]["full_name"] == "גדולה")
db.set_need_weights({"income": 100, "souls": 0, "money": 0,
                     "recency": 0, "housing": 0, "medical": 0, "children": 0})
check("income-weighted → low income first", db.get_one_time_list()[0]["full_name"] == "עניה")
db.set_need_weights(db.DEFAULT_NEED_WEIGHTS)   # restore default


# ══════════════════════════════════════════════════
# סיכום
# ══════════════════════════════════════════════════
print()
print("=" * 55)
total = _passed + len(_errors)
print(f"עברו: {_passed}/{total}  |  נכשלו: {len(_errors)}")
if _errors:
    print("\nנכשלו:")
    for e in _errors:
        print(f"  ✗ {e}")
else:
    print("✓ כל הבדיקות עברו בהצלחה!")

try: os.unlink(_TMP_DB)
except: pass

if _errors:
    sys.exit(1)
