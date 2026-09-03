# -*- coding: utf-8 -*-
"""לשונית 'צינתוקים' (v2.81) — הודעה קולית אוטומטית לזכאי החלוקה של השבוע
דרך ימות המשיח, מתוך התוכנה.

הזרימה (לפי מסמך ההדמיה שאישר המשתמש): רשימת הזכאים של מסך "חלוקה ורישום"
→ סינון חריגים (בלי מספר / מספר שבור / מספר כפול) → בדיקה למספר של המנהל →
אישור מפורש → שליחה עם מעקב חי (RunCampaign + GetCampaignStatus) → תוצאות
עם "שלח שוב לנכשלים" → היסטוריה מסונכרנת בין שני המחשבים.

v2.88: כל המספרים של כל מקבל מצולצלים (#gaira, בלי קומבו בחירה); אפשר לטעון
רשימה מחלוקה קודמת (#9hgvi, קליק ימני בלשונית "חלוקות קודמות"); ההודעה
מושמעת למי שברשימה כשהוא מתקשר חזרה לקו (#z4xy9, פרטי); אין חסימת שעות
(#n6wte). v2.89: פרסום בשלוחה 1 (#kx6wd) הוא ידני בלבד — כפתור עם אזהרה;
שום דבר לא מתפרסם אוטומטית (הודעת מקבלים היא פרטית — הכרעת המשתמש).

הגנות: שומר שליחה-כפולה לאותה חלוקה (חוצה-מחשבים, דרך הסנכרון),
נעילת הכפתור בזמן שליחה."""
import json
import os
import time
from datetime import datetime, timedelta, timezone

from PyQt6.QtCore import Qt, QThread, pyqtSignal, QDateTime, QTimer, QEventLoop
from PyQt6.QtGui import QColor, QIcon
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QComboBox, QMessageBox, QProgressBar, QScrollArea, QDialog, QLineEdit,
    QListWidget, QListWidgetItem, QFileDialog, QInputDialog, QTextEdit,
    QDateTimeEdit, QProgressDialog, QSizePolicy
)

import database as db
from utils import call_history, timefmt, tts, yemot
from utils.ui import busy_cursor, enable_touch_scroll, line_icon
# The design language (cards, glossy buttons, chips, page background) is shared
# with the main "חלוקה ורישום" screen so both read as one app.
from tabs.group_update import (_BG, _CARD_QSS, _CHIP_QSS, _CHIP_GREEN, _CHIP_AMBER,
                               _LBL, _BTN_PRIMARY, _BTN_GHOST, _BTN_ACCENT, _BTN_PRINT,
                               _step_badge, _step_card, _metric, _set_metric)
# A quiet text-only button for a rare, sensitive action (publishing to ext. 1).
_BTN_LINK = ("QPushButton{background:transparent; color:#b45309; border:none;"
             " font-weight:700; font-size:13px; padding:0 8px; min-height:38px;"
             " text-decoration:underline;}"
             "QPushButton:hover{color:#92400e;}")


class _PollWorker(QThread):
    """Polls the campaign status off the UI thread every few seconds until the
    campaign finishes (or ~15 minutes pass)."""
    tick = pyqtSignal(object)          # status dict | Exception

    def __init__(self, campaign_id: str, parent=None, since_iso: str = ""):
        super().__init__(parent)
        self.campaign_id = campaign_id
        self.since_iso = since_iso        # campaign send time — survey answers before it don't count
        self._rows = None                 # last survey rows fetched (v3.02)
        self._n = 0
        self._stop = False

    def stop(self):
        self._stop = True

    def _merge_survey(self, st: dict, force: bool = False):
        """v3.02 — every ~5th tick (and at the end) read the survey extension's
        answers and stamp them onto the entries; a fetch failure keeps the
        last rows (the campaign status itself is what matters live)."""
        self._n += 1
        if force or self._n % 5 == 1:
            try:
                self._rows = yemot.fetch_survey_rows()
            except Exception:
                pass
        if self._rows is not None:
            yemot.merge_survey_answers(st.get("entries") or [], self._rows, self.since_iso)
            st["answers"] = yemot.answer_counts(st.get("entries") or [])

    def run(self):
        errors = 0
        for _ in range(225):
            if self._stop:
                return
            try:
                st = yemot.get_campaign_status(self.campaign_id)
            except Exception as e:
                errors += 1
                if errors >= 5:
                    self.tick.emit(e)
                    return
                time.sleep(8)
                continue
            errors = 0
            self._merge_survey(st, force=bool(st.get("finished")))
            self.tick.emit(st)
            if st.get("finished"):
                return
            for _ in range(8):          # 4s in small slices → stop() is snappy
                if self._stop:
                    return
                time.sleep(0.5)


class _CallbackWorker(QThread):
    """v2.96 — live watch after a CLASSIC tzintuk: Yemot keeps no call
    history, so while the window is open we poll the line's LIVE calls
    (GetIncomingCalls) and mark every target number that calls back — and
    (v3.02) what they answered on the survey extension (1/2/3)."""
    tick = pyqtSignal(object)          # snapshot dict | Exception

    def __init__(self, targets: dict, deadline: float,
                 seed_entries=None, parent=None, since_iso: str = ""):
        super().__init__(parent)
        self.tracker = yemot.CallbackTracker(targets)
        if seed_entries:
            self.tracker.seed(seed_entries)
        self.deadline = float(deadline)
        self.since_iso = since_iso
        self._rows = None
        self._rows_at = 0.0
        self._stop = False

    def stop(self):
        self._stop = True

    def extend(self, seconds: int):
        self.deadline += seconds

    def _snapshot(self, done: bool, changed: bool = False, error: str = ""):
        returned, _ = self.tracker.counts()
        entries = self.tracker.entries()
        # Survey answers: refreshed once a minute (and when the window closes).
        if self._rows is None or done or time.time() - self._rows_at > 60:
            try:
                self._rows = yemot.fetch_survey_rows()
                self._rows_at = time.time()
            except Exception:
                pass
        if self._rows is not None:
            _, ans_changed = yemot.merge_survey_answers(entries, self._rows, self.since_iso)
            changed = changed or ans_changed
        return {"returned": returned, "answers": yemot.answer_counts(entries),
                "entries": entries, "changed": changed,
                "remaining": max(0.0, self.deadline - time.time()),
                "done": done, "error": error}

    def run(self):
        errors = 0
        self.tick.emit(self._snapshot(False))
        while not self._stop and time.time() < self.deadline:
            try:
                calls = yemot.get_incoming_calls()
                changed = self.tracker.update(
                    calls, datetime.now(timezone.utc).isoformat())
                errors = 0
                self.tick.emit(self._snapshot(False, changed))
            except Exception as e:
                errors += 1
                if errors in (3, 20):    # tell the UI, but keep trying
                    self.tick.emit(self._snapshot(False, error=str(e)))
                for _ in range(20):
                    if self._stop:
                        break
                    time.sleep(0.5)
                continue
            for _ in range(10):          # ~5s in slices → stop() is snappy
                if self._stop:
                    break
                time.sleep(0.5)
        self.tick.emit(self._snapshot(True))


class _AddPersonDialog(QDialog):
    """Pick an active recipient who is not already on the call list."""

    def __init__(self, exclude_ids, parent=None):
        super().__init__(parent)
        self.setWindowTitle("הוסף אדם לרשימת הצינתוקים")
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.resize(420, 480)
        self.picked = None
        self._all = [r for r in db.get_all_recipients("פעיל")
                     if r["id"] not in exclude_ids]
        lay = QVBoxLayout(self)
        self.search = QLineEdit()
        self.search.setPlaceholderText("חיפוש לפי שם / טלפון…")
        self.search.textChanged.connect(self._filter)
        lay.addWidget(self.search)
        self.listw = QListWidget()
        self.listw.itemDoubleClicked.connect(lambda _i: self._accept())
        lay.addWidget(self.listw, 1)
        btns = QHBoxLayout()
        ok = QPushButton("הוסף")
        ok.setObjectName("primary")
        ok.clicked.connect(self._accept)
        cancel = QPushButton("ביטול")
        cancel.clicked.connect(self.reject)
        btns.addWidget(ok)
        btns.addWidget(cancel)
        btns.addStretch()
        lay.addLayout(btns)
        self._filter()

    def _filter(self):
        text = self.search.text().strip()
        rows = db.filter_recipients(self._all, text, limit=300) if text else self._all[:300]
        self.listw.clear()
        for r in rows:
            phones = yemot.pick_phones(dict(r))
            label = r["full_name"] + ("  ·  " + phones[0] if phones else "  ·  (אין מספר)")
            it = QListWidgetItem(label)
            it.setData(Qt.ItemDataRole.UserRole, dict(r))
            self.listw.addItem(it)

    def _accept(self):
        it = self.listw.currentItem()
        if it is None:
            return
        self.picked = it.data(Qt.ItemDataRole.UserRole)
        self.accept()


class _FreeListDialog(QDialog):
    """רשימה עצמאית (#1/9) — הדבקת מספרי טלפון או טעינת קובץ אקסל, בלי שום
    קשר לרשימות החלוקה. מציג בזמן-אמת כמה מספרים תקינים זוהו, כדי שהמפעיל
    יידע ב-100% מה עומד להישלח."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("רשימה עצמאית לצינתוק")
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.resize(520, 520)
        self.entries = []              # [(phone, name), …] — the parsed result
        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        info = QLabel("הדבק כאן מספרי טלפון (מספר בכל שורה, אפשר גם שם ליד "
                      "המספר) — או טען קובץ אקסל. הרשימה הזו נפרדת לגמרי "
                      "מרשימות החלוקה.")
        info.setWordWrap(True)
        lay.addWidget(info)
        self.text = QTextEdit()
        self.text.setPlaceholderText("לדוגמה:\n0501234567\tמשפחת כהן\n0521111222")
        self.text.setAcceptRichText(False)
        self.text.textChanged.connect(self._reparse)
        lay.addWidget(self.text, 1)
        btn_xls = QPushButton("📊 טען מקובץ אקסל…")
        btn_xls.setObjectName("neutral")
        btn_xls.setToolTip("קורא את כל המספרים מהקובץ ומציג אותם כאן — "
                           "מה שרואים זה מה שנשלח")
        btn_xls.clicked.connect(self._load_excel)
        lay.addWidget(btn_xls)
        self.lbl_count = QLabel("")
        self.lbl_count.setObjectName("subtitle")
        self.lbl_count.setWordWrap(True)
        lay.addWidget(self.lbl_count)
        btns = QHBoxLayout()
        self.btn_ok = QPushButton("הצג את הרשימה »")
        self.btn_ok.setObjectName("primary")
        self.btn_ok.clicked.connect(self._accept)
        cancel = QPushButton("ביטול")
        cancel.clicked.connect(self.reject)
        btns.addWidget(self.btn_ok)
        btns.addWidget(cancel)
        btns.addStretch()
        lay.addLayout(btns)
        self._reparse()

    @staticmethod
    def _parse_text(raw: str):
        """[(phone, name)], bad_tokens — a token is a phone when it normalizes;
        the rest of its line becomes the name."""
        entries, bad = [], []
        for line in raw.splitlines():
            line = line.strip()
            if not line:
                continue
            # Numbers may contain spaces ("050 123 4567") and Excel-style
            # numbers may lack the leading zero — yemot.find_phones handles
            # both; the words left over are the name.
            phones, name, bad_toks = yemot.find_phones(
                line.replace(";", " ").replace(",", " ").replace("\t", " "))
            bad.extend(bad_toks)
            for p in phones:
                entries.append((p, name))
        return entries, bad

    def _reparse(self):
        self.entries, bad = self._parse_text(self.text.toPlainText())
        seen, unique = set(), []
        for p, n in self.entries:
            if p not in seen:
                seen.add(p)
                unique.append((p, n))
        self.entries = unique
        msg = f"זוהו {len(self.entries)} מספרים תקינים."
        if bad:
            msg += f"  ⚠ {len(bad)} קטעים נראים כמו מספר אבל אינם תקינים: " \
                   + ", ".join(bad[:5]) + ("…" if len(bad) > 5 else "")
        self.lbl_count.setText(msg)
        self.btn_ok.setEnabled(bool(self.entries))

    def _load_excel(self):
        path, _f = QFileDialog.getOpenFileName(
            self, "בחר קובץ אקסל", "", "קובצי אקסל (*.xlsx *.xlsm);;כל הקבצים (*.*)")
        if not path:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    phones, words = [], []
                    for cell in row:
                        if cell is None:
                            continue
                        # A number-typed cell lost its leading zero / gained
                        # ".0" — normalize_phone_loose restores both.
                        if isinstance(cell, float) and cell.is_integer():
                            cell = int(cell)
                        val = str(cell).strip()
                        if not val:
                            continue
                        p = yemot.normalize_phone_loose(val)
                        if p:
                            phones.append(p)
                        elif not any(ch.isdigit() for ch in val):
                            words.append(val)
                    for p in phones:
                        lines.append(f"{p}\t{' '.join(words)}".rstrip())
            wb.close()
        except Exception as e:
            QMessageBox.warning(self, "טעינת אקסל", f"קריאת הקובץ נכשלה:\n{e}")
            return
        if not lines:
            QMessageBox.information(self, "טעינת אקסל",
                                    "לא נמצאו מספרי טלפון תקינים בקובץ.")
            return
        # ההדבקה לתיבה — כך המפעיל רואה בעיניים בדיוק מה נטען מהאקסל.
        self.text.setPlainText("\n".join(lines))

    def _accept(self):
        if not self.entries:
            return
        self.accept()


class _SendModeDialog(QDialog):
    """אישור שליחה + בחירת סוג הצינתוק (#1/9): רגיל (משמיע הודעה למי שעונה)
    או קלאסי (צלצול קצר בלי מענה — מי שמתקשר חזרה שומע את ההודעה)."""

    def __init__(self, summary: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("אישור שליחה")
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.mode = None               # 'voice' | 'classic'
        from PyQt6.QtWidgets import QRadioButton
        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        lbl = QLabel(summary)
        lbl.setWordWrap(True)
        lay.addWidget(lbl)
        self.rb_voice = QRadioButton("צינתוק עם הודעה (רגיל) — מי שעונה שומע "
                                     "את ההודעה מיד")
        self.rb_voice.setChecked(True)
        lay.addWidget(self.rb_voice)
        self.rb_classic = QRadioButton("צינתוק קלאסי — צלצול קצר בלי מענה "
                                       "(כמעט בלי עלות)")
        lay.addWidget(self.rb_classic)
        note = QLabel("בצינתוק קלאסי הטלפון מצלצל ומתנתק — אי אפשר לענות "
                      "לשיחה. מי שרואה שיחה שלא נענתה ומתקשר חזרה לקו — "
                      "שומע את ההודעה. אחרי השליחה התוכנה עוקבת כחצי שעה "
                      "בזמן אמת מי חזר לשיחה ומי אישר הגעה בהקשת 7.")
        note.setObjectName("subtitle")
        note.setWordWrap(True)
        lay.addWidget(note)
        btns = QHBoxLayout()
        ok = QPushButton("שלח עכשיו »")
        ok.setObjectName("primary")
        ok.clicked.connect(self._accept)
        cancel = QPushButton("ביטול")
        cancel.clicked.connect(self.reject)
        btns.addWidget(ok)
        btns.addWidget(cancel)
        btns.addStretch()
        lay.addLayout(btns)

    def _accept(self):
        self.mode = "classic" if self.rb_classic.isChecked() else "voice"
        self.accept()


class _TestStatusDialog(QDialog):
    """מעקב חי אחרי שיחת הבדיקה — כשמספר הבדיקה אינו הטלפון שביד המפעיל
    רואים כאן אם השיחה נענתה, נכשלה או שהוקש 7, בלי לנחש."""

    def __init__(self, campaign_id: str, phone: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("שליחת בדיקה — מעקב")
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.setMinimumWidth(400)
        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        head = QLabel(f"📞 מצלצל אל {phone}")
        head.setStyleSheet(_LBL + " font-weight:bold;")
        lay.addWidget(head)
        self.lbl_status = QLabel("⏳ ממתין לעדכון מהשרת…")
        self.lbl_status.setWordWrap(True)
        self.lbl_status.setStyleSheet(_LBL + " font-size:15px;")
        lay.addWidget(self.lbl_status)
        note = QLabel("הסטטוס מתעדכן כאן אוטומטית כל כמה שניות. "
                      "אפשר לסגור את החלון — השיחה תמשיך כרגיל.")
        note.setObjectName("subtitle")
        note.setWordWrap(True)
        lay.addWidget(note)
        btns = QHBoxLayout()
        self.btn_close = QPushButton("סגור")
        self.btn_close.clicked.connect(self.accept)
        btns.addWidget(self.btn_close)
        btns.addStretch()
        lay.addLayout(btns)
        self._worker = _PollWorker(campaign_id, self,
                                   datetime.now(timezone.utc).isoformat())
        self._worker.tick.connect(self._on_tick)
        self._worker.start()

    def _on_tick(self, st):
        if isinstance(st, Exception):
            self.lbl_status.setText(
                "⚠ המעקב נכשל (בעיית תקשורת) — השיחה עצמה יוצאת כרגיל מהשרת "
                "של ימות.")
            self.lbl_status.setStyleSheet(_LBL + " font-size:15px; color:#a35b00;")
            return
        entries = st.get("entries") or []
        e = entries[0] if entries else {}
        status = e.get("status") or ""
        ans = str(e.get("answer") or "")
        if ans in yemot.ANSWER_KEYS:
            txt = f"✓ השיחה נענתה — ובסקר הוקש {ans} ({yemot.answer_label(ans)})"
            color = "#166534"
        elif e.get("ok"):
            txt = ("✓ השיחה נענתה ע\"י תא קולי — ההודעה הושארה בו"
                   if status == "amd" else "✓ השיחה נענתה — ההודעה הושמעה")
            color = "#0f6e56"
        elif e.get("failed"):
            txt = {"no_answer": "לא היה מענה לשיחה",
                   "busy": "הקו תפוס"}.get(status, f"השיחה נכשלה ({status})")
            txt, color = "✗ " + txt, "#a32d2d"
        else:
            txt, color = "📞 מצלצל עכשיו…", "#0f4c81"
        # עוצרים רק כשהקמפיין באמת הסתיים — אם נעצור כבר על "נענתה" נפספס
        # את הסטטוס הסופי של השיחה.
        if st.get("finished"):
            self._worker.stop()
        else:
            txt += "\n(ממשיך לעקוב…)"
        self.lbl_status.setText(txt)
        self.lbl_status.setStyleSheet(
            _LBL + f" font-size:15px; font-weight:bold; color:{color};")

    def closeEvent(self, ev):
        if self._worker is not None:
            self._worker.stop()
        super().closeEvent(ev)

    def accept(self):
        if self._worker is not None:
            self._worker.stop()
        super().accept()


class _ScheduleDialog(QDialog):
    """בחירת תאריך ושעה לצינתוק מתוזמן (#xi85i) — התזמון נשמר בשרת של ימות
    ורץ גם כשהמחשב כבוי."""

    def __init__(self, count: int, parent=None, smart_hint: str = ""):
        super().__init__(parent)
        self.setWindowTitle("תזמון שליחה")
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.when = None
        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        info = QLabel(f"הצינתוק יישלח ל-{count} נמענים במועד שתבחר.\n"
                      "השליחה מתבצעת מהשרת של ימות המשיח — "
                      "המחשב לא חייב להיות דלוק באותה שעה.")
        info.setWordWrap(True)
        lay.addWidget(info)
        self.dt_edit = QDateTimeEdit()
        self.dt_edit.setCalendarPopup(True)
        # RTL reverses the date/time sections (widgets.DateEdit trap): go LTR
        # FIRST and only then set the display format, or the section order
        # sticks reversed. The dialog around the field stays RTL.
        self.dt_edit.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        self.dt_edit.setDisplayFormat("dd/MM/yyyy  HH:mm")
        self.dt_edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tomorrow = datetime.now() + timedelta(days=1)
        self.dt_edit.setDateTime(QDateTime(
            tomorrow.year, tomorrow.month, tomorrow.day, 9, 0))
        self.dt_edit.setMinimumDateTime(QDateTime.currentDateTime())
        lay.addWidget(self.dt_edit)
        if smart_hint:              # #y7jr0 שלב 1 — המלצה מתוך ההיסטוריה
            hint = QLabel("💡 " + smart_hint)
            hint.setObjectName("subtitle")
            hint.setWordWrap(True)
            lay.addWidget(hint)
        btns = QHBoxLayout()
        ok = QPushButton("תזמן »")
        ok.setObjectName("primary")
        ok.clicked.connect(self._accept)
        cancel = QPushButton("ביטול")
        cancel.clicked.connect(self.reject)
        btns.addWidget(ok)
        btns.addWidget(cancel)
        btns.addStretch()
        lay.addLayout(btns)

    def _accept(self):
        when = self.dt_edit.dateTime().toPyDateTime().replace(second=0,
                                                              microsecond=0)
        if when <= datetime.now() + timedelta(minutes=2):
            QMessageBox.warning(self, "תזמון שליחה",
                                "בחר מועד עתידי (לפחות כמה דקות מעכשיו).")
            return
        self.when = when
        self.accept()


class _TaskWorker(QThread):
    """Runs one blocking callable off the UI thread (TTS synthesis takes a few
    seconds — freezing the dialog would look like a hang)."""
    done = pyqtSignal(object)          # result | Exception

    def __init__(self, fn, parent=None):
        super().__init__(parent)
        self._fn = fn

    def run(self):
        try:
            self.done.emit(self._fn())
        except Exception as e:
            self.done.emit(e)


class TtsDialog(QDialog):
    """יצירת הקלטת צינתוק מטקסט (#9vy1b) — כותבים את ההודעה, בוחרים קול,
    מאזינים לדוגמה, וההקלטה נשמרת במאגר ומועלית לימות."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("יצירת הקלטה מטקסט")
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.resize(560, 460)
        self.result_path = ""          # ההקלטה המאושרת (wav/mp3)
        self.result_text = ""
        self.result_voice = ""
        self.result_note = ""          # הערת נפילה-לקול-גיבוי, אם הייתה
        self._worker = None
        self._preview_key = None       # (text, voice, rate) של הקובץ האחרון שנוצר
        self._preview_base = os.path.join(tts.recordings_dir(), "_preview")
        self._preview_path = ""        # הנתיב שנכתב בפועל (הסיומת לפי המנוע)

        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        intro = QLabel("כתוב את נוסח ההודעה, והמחשב יקריא אותה בקול טבעי (חינם, "
                       "דרך האינטרנט). טיפ: פסיקים ונקודות יוצרים הפסקות טבעיות, "
                       "ואפשר להוסיף ניקוד למילה שנשמעת לא נכון.")
        intro.setObjectName("subtitle")
        intro.setWordWrap(True)
        lay.addWidget(intro)

        self.text = QTextEdit()
        self.text.setAcceptRichText(False)
        self.text.setPlaceholderText(tts.DEFAULT_TEXT)
        last = (db.get_setting(tts.SET_LAST_TEXT) or "").strip()
        self.text.setPlainText(last or tts.DEFAULT_TEXT)
        lay.addWidget(self.text, 1)

        opts = QHBoxLayout()
        opts.addWidget(QLabel("קול:"))
        self.voice = QComboBox()
        for label, _vid in tts.VOICES:
            self.voice.addItem(label)
        last_v = db.get_setting(tts.SET_LAST_VOICE) or ""
        for i, (_l, vid) in enumerate(tts.VOICES):
            if vid == last_v:
                self.voice.setCurrentIndex(i)
        opts.addWidget(self.voice)
        opts.addSpacing(12)
        opts.addWidget(QLabel("מהירות:"))
        self.rate = QComboBox()
        for label, _r in tts.RATES:
            self.rate.addItem(label)
        opts.addWidget(self.rate)
        opts.addStretch()
        lay.addLayout(opts)

        self.status = QLabel("")
        self.status.setObjectName("subtitle")
        self.status.setWordWrap(True)
        lay.addWidget(self.status)

        btns = QHBoxLayout()
        self.btn_preview = QPushButton("🔊 השמע לי לדוגמה")
        self.btn_preview.setObjectName("neutral")
        self.btn_preview.setToolTip("יוצר את ההקלטה ופותח אותה בנגן של המחשב — "
                                    "לפני שמאשרים")
        self.btn_preview.clicked.connect(self._preview)
        btns.addWidget(self.btn_preview)
        btns.addStretch()
        self.btn_ok = QPushButton("אשר — שמור במאגר והעלה לצינתוק")
        self.btn_ok.setObjectName("primary")
        self.btn_ok.clicked.connect(self._accept)
        btns.addWidget(self.btn_ok)
        cancel = QPushButton("ביטול")
        cancel.clicked.connect(self.reject)
        btns.addWidget(cancel)
        lay.addLayout(btns)

    # ── generation ────────────────────────────────────────────────────────────

    def _params(self):
        text = self.text.toPlainText().strip()
        voice = tts.VOICES[self.voice.currentIndex()][1]
        rate = tts.RATES[self.rate.currentIndex()][1]
        return text, voice, rate

    def _generate(self, on_ready):
        """יוצר (ברקע) את ההקלטה לקובץ ה-preview; on_ready() נקרא בהצלחה."""
        text, voice, rate = self._params()
        if not text:
            QMessageBox.warning(self, "יצירת הקלטה", "כתוב קודם את נוסח ההודעה.")
            return
        if self._worker is not None:
            return
        key = (text, voice, rate)
        if (key == self._preview_key and self._preview_path
                and os.path.exists(self._preview_path)):
            on_ready()                    # כבר נוצר בדיוק אותו דבר — אין צורך שוב
            return
        self._set_busy(True)
        self._worker = _TaskWorker(
            lambda: tts.synthesize(text, voice, self._preview_base, rate), self)

        def _done(res):
            self._worker = None
            self._set_busy(False)
            if isinstance(res, Exception):
                QMessageBox.warning(self, "יצירת הקלטה", str(res))
                return
            self._preview_path, note = res
            self._preview_key = key
            if note:
                self.status.setText("⚠ " + note)
            on_ready()
        self._worker.done.connect(_done)
        self._worker.start()

    def _set_busy(self, busy: bool):
        for b in (self.btn_preview, self.btn_ok):
            b.setEnabled(not busy)
        self.status.setText("🎙 יוצר את ההקלטה… (כמה שניות)" if busy else "")

    def _preview(self):
        def _play():
            try:
                os.startfile(self._preview_path)
                note = self.status.text()
                self.status.setText(
                    (note + "\n" if note.startswith("⚠") else "")
                    + "ההקלטה נפתחה בנגן — אם הנוסח טוב, לחץ \"אשר\" למטה.")
            except OSError as e:
                QMessageBox.warning(self, "השמעה", f"פתיחת הנגן נכשלה: {e}")
        self._generate(_play)

    def _accept(self):
        def _finish():
            text, voice, _rate = self._params()
            self.result_path = self._preview_path
            self.result_text = text
            self.result_voice = voice
            self.result_note = (self.status.text()
                                if self.status.text().startswith("⚠") else "")
            db.set_setting(tts.SET_LAST_TEXT, text)
            db.set_setting(tts.SET_LAST_VOICE, voice)
            self.accept()
        self._generate(_finish)

    def closeEvent(self, ev):
        if self._worker is not None:
            self._worker.done.disconnect()
        super().closeEvent(ev)


class LibraryDialog(QDialog):
    """מאגר ההקלטות (#kgmcw) — כל הקלטה שנוצרה מטקסט או הועלתה נשמרת כאן,
    ואפשר להשמיע / להעלות שוב לצינתוק בלי להקליט מחדש."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("מאגר הקלטות")
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.resize(640, 440)
        self.picked_path = ""          # ההקלטה שנבחרה להעלאה לצינתוק
        self.picked_name = ""

        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        intro = QLabel("כל הקלטה שיצרת מטקסט או העלית מקובץ נשמרת כאן במחשב, "
                       "כדי שלא תצטרך להקליט מחדש. לחיצה כפולה = השמעה.")
        intro.setObjectName("subtitle")
        intro.setWordWrap(True)
        lay.addWidget(intro)

        self.listw = QListWidget()
        self.listw.itemDoubleClicked.connect(lambda _i: self._play())
        lay.addWidget(self.listw, 1)

        tools = QHBoxLayout()
        b_play = QPushButton("🔊 השמע")
        b_play.setObjectName("neutral")
        b_play.clicked.connect(self._play)
        tools.addWidget(b_play)
        b_rename = QPushButton("שנה שם")
        b_rename.setObjectName("neutral")
        b_rename.clicked.connect(self._rename)
        tools.addWidget(b_rename)
        b_del = QPushButton("מחק")
        b_del.setObjectName("danger")
        b_del.clicked.connect(self._delete)
        tools.addWidget(b_del)
        tools.addStretch()
        lay.addLayout(tools)

        btns = QHBoxLayout()
        btns.addStretch()
        self.b_use = QPushButton("השתמש בהקלטה זו »")
        self.b_use.setToolTip("מעלה את ההקלטה הנבחרת לימות — היא שתושמע בצינתוק")
        self.b_use.setObjectName("primary")
        self.b_use.clicked.connect(self._use)
        btns.addWidget(self.b_use)
        close = QPushButton("סגור")
        close.clicked.connect(self.reject)
        btns.addWidget(close)
        lay.addLayout(btns)
        self._refresh()

    def _refresh(self):
        self.listw.clear()
        items = tts.library_list()
        for it in items:
            when = (it.get("created") or "")[:16]
            src = "נוצרה מטקסט" if it.get("source") == "tts" else "קובץ שהועלה"
            li = QListWidgetItem(f"🎵 {it.get('name')}   ·   {when}   ·   {src}")
            if it.get("text"):
                li.setToolTip(it["text"])
            li.setData(Qt.ItemDataRole.UserRole, it)
            self.listw.addItem(li)
        empty = not items
        self.b_use.setEnabled(not empty)
        if empty:
            self.listw.addItem(QListWidgetItem(
                "המאגר ריק — צור הקלטה מטקסט או העלה קובץ, והיא תישמר כאן."))

    def _current(self):
        it = self.listw.currentItem()
        return it.data(Qt.ItemDataRole.UserRole) if it else None

    def _play(self):
        item = self._current()
        if not item:
            return
        try:
            os.startfile(tts.library_path(item))
        except OSError as e:
            QMessageBox.warning(self, "השמעה", f"פתיחת הנגן נכשלה: {e}")

    def _rename(self):
        item = self._current()
        if not item:
            return
        name, ok = QInputDialog.getText(self, "שנה שם", "שם ההקלטה:",
                                        text=item.get("name") or "")
        if ok and name.strip():
            tts.library_rename(item["id"], name.strip())
            self._refresh()

    def _delete(self):
        item = self._current()
        if not item:
            return
        ans = QMessageBox.question(
            self, "מחיקת הקלטה",
            f"למחוק את ההקלטה \"{item.get('name')}\" מהמאגר?\n"
            "(הקלטה שכבר הועלתה לימות ממשיכה לפעול שם)",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if ans == QMessageBox.StandardButton.Yes:
            tts.library_delete(item["id"])
            self._refresh()

    def _use(self):
        item = self._current()
        if not item:
            QMessageBox.information(self, "מאגר הקלטות", "בחר קודם הקלטה מהרשימה.")
            return
        self.picked_path = tts.library_path(item)
        self.picked_name = item.get("name") or ""
        self.accept()


class TzintukimTab(QWidget):
    """מסך הצינתוקים — ראו docstring של המודול."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.main = parent
        self._rows = []           # [{'rec', 'phones', 'send', 'checked', 'why'}]
        self._worker = None
        self._cb_worker = None    # v2.96: classic-tzintuk callback watcher
        self._cb_last_persist = 0.0
        self._active_guid = ""    # DB guid of the campaign being tracked
        self._last_failed = []    # [{'phone','name'}] from the last finished run
        self._last_entries = []   # per-number results shown in the table (survive refresh)
        self._sched_checker = None   # worker probing the server for due schedules
        self._batch = None        # #9hgvi: past-distribution batch loaded as list
        self._free = None         # #1/9: standalone list [(phone, name), …]
        self._stats = {}          # #y7jr0: per-phone answer history (tooltips)
        # #ifc70 — the week list is NOT loaded automatically; the operator loads
        # it with an explicit button (a past batch via #9hgvi counts as loaded).
        self._list_loaded = False
        self._build_ui()
        # Due schedules / an interrupted campaign are picked up shortly after
        # launch even if nobody opens this tab (the message promised "results
        # on the next start").
        QTimer.singleShot(6000, self._maybe_resume_tracking)

    # ── UI ────────────────────────────────────────────────────────────────────
    #
    # v3.00 — the screen was rebuilt from scratch as ONE clear flow:
    #   header (title + connection chip)
    #   ① נמענים  — where the list comes from, the counters, the table
    #   ② ההודעה — what the phones will play (create / upload / library / test)
    #   היסטוריה — every past send
    #   sticky bottom bar — ③ שליחה: summary, schedule, the big send button,
    #                        and the live-progress / scheduled strips.
    # Every widget the logic talks to keeps its old attribute name.

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        surface = QWidget()
        surface.setObjectName("tz-surface")
        surface.setStyleSheet(f"QWidget#tz-surface{{background:{_BG};}}")
        root.addWidget(surface, 1)
        s_lay = QVBoxLayout(surface)
        s_lay.setContentsMargins(0, 0, 0, 0)
        s_lay.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea{background:transparent;}"
                             "QScrollArea>QWidget>QWidget{background:transparent;}")
        enable_touch_scroll(scroll)
        content = QWidget()
        scroll.setWidget(content)
        lay = QVBoxLayout(content)
        lay.setSpacing(12)
        lay.setContentsMargins(20, 12, 20, 8)
        s_lay.addWidget(scroll, 1)

        # ── Header: title · subtitle · connection chip ────────────────────────
        head = QHBoxLayout()
        head.setSpacing(12)
        title = QLabel("צינתוקים")
        title.setStyleSheet("color:#064e3b; font-size:22px; font-weight:800; " + _LBL)
        head.addWidget(title)
        sub = QLabel("הודעה קולית לזכאי החלוקה דרך ימות המשיח")
        sub.setStyleSheet("color:#64748b; font-size:13px; " + _LBL)
        head.addWidget(sub)
        head.addStretch()
        # Connected chip (shown when credentials exist) …
        self.lbl_ok = QLabel("●  מחובר לימות המשיח")
        self.lbl_ok.setStyleSheet(_CHIP_GREEN)
        head.addWidget(self.lbl_ok)
        # … or the not-configured chip + a way in (hidden once configured).
        self.banner = QWidget()
        b_lay = QHBoxLayout(self.banner)
        b_lay.setContentsMargins(0, 0, 0, 0)
        b_lay.setSpacing(8)
        b_txt = QLabel("●  עוד לא חובר לימות המשיח")
        b_txt.setStyleSheet(_CHIP_AMBER)
        b_lay.addWidget(b_txt)
        b_btn = QPushButton("פתח הגדרות")
        b_btn.setStyleSheet(_BTN_GHOST)
        b_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        b_btn.clicked.connect(self._goto_settings)
        b_lay.addWidget(b_btn)
        head.addWidget(self.banner)
        lay.addLayout(head)

        # ── ① נמענים ─────────────────────────────────────────────────────────
        card, c_lay, c_head = _step_card(
            "1", "נמענים", "מי יקבל את הצינתוק")
        self.btn_load = QPushButton("  רשימת החלוקה הנוכחית")
        self.btn_load.setStyleSheet(_BTN_PRIMARY)
        self.btn_load.setIcon(QIcon(line_icon("import", 18, "#ffffff")))
        self.btn_load.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_load.setToolTip("טוען לכאן את רשימת הזכאים ממסך \"חלוקה ורישום\" "
                                 "(בלי הרזרבות)")
        self.btn_load.clicked.connect(self._load_week_list)
        c_head.addWidget(self.btn_load)
        self.btn_free = QPushButton("  רשימה עצמאית…")
        self.btn_free.setStyleSheet(_BTN_GHOST)
        self.btn_free.setIcon(QIcon(line_icon("doc", 18, "#475569")))
        self.btn_free.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_free.setToolTip("מדביקים מספרי טלפון או בוחרים קובץ אקסל — "
                                 "בלי שום קשר לרשימות החלוקה")
        self.btn_free.clicked.connect(self._load_free_list)
        c_head.addWidget(self.btn_free)

        # Empty state — until the operator picks a source.
        self.load_frame = QLabel(
            "עוד לא נטענה רשימה. בחר למעלה: רשימת החלוקה הנוכחית או רשימה "
            "עצמאית. חלוקה קודמת — קליק ימני בלשונית \"חלוקות קודמות\".")
        self.load_frame.setStyleSheet("color:#94a3b8; font-size:13px; padding:26px 0; " + _LBL)
        self.load_frame.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.load_frame.setWordWrap(True)
        c_lay.addWidget(self.load_frame)

        self.list_frame = QWidget()
        lf_lay = QVBoxLayout(self.list_frame)
        lf_lay.setContentsMargins(0, 0, 0, 0)
        lf_lay.setSpacing(8)

        # Source strip — only for a past distribution / a standalone list.
        self.batch_frame = QFrame()
        self.batch_frame.setStyleSheet("QFrame{background:#eef2ff; border:1px solid"
                                       " #c7d2fe; border-radius:9px;}")
        bt_lay = QHBoxLayout(self.batch_frame)
        bt_lay.setContentsMargins(12, 5, 12, 5)
        self.lbl_batch = QLabel("")
        self.lbl_batch.setStyleSheet("color:#3730a3; font-weight:600; " + _LBL)
        self.lbl_batch.setWordWrap(True)
        bt_lay.addWidget(self.lbl_batch, 1)
        self.btn_back = QPushButton("חזור לרשימת השבוע")
        self.btn_back.setStyleSheet(_BTN_GHOST)
        self.btn_back.clicked.connect(self._clear_batch)
        bt_lay.addWidget(self.btn_back)
        self.batch_frame.setVisible(False)
        lf_lay.addWidget(self.batch_frame)

        # Counters + list tools on one row.
        tools = QHBoxLayout()
        tools.setSpacing(8)
        self.m_total = _metric("זכאים", _CHIP_QSS)
        self.m_ready = _metric("מוכנים לשליחה", _CHIP_GREEN)
        self.m_bad = _metric("חריגים", _CHIP_AMBER)
        for m in (self.m_total, self.m_ready, self.m_bad):
            tools.addWidget(m["frame"])
        tools.addStretch()
        btn_check_all = QPushButton("סמן את כולם")
        btn_check_all.setStyleSheet(_BTN_GHOST)
        btn_check_all.clicked.connect(lambda: self._set_all_checked(True))
        tools.addWidget(btn_check_all)
        btn_uncheck_all = QPushButton("נקה סימון")
        btn_uncheck_all.setStyleSheet(_BTN_GHOST)
        btn_uncheck_all.clicked.connect(lambda: self._set_all_checked(False))
        tools.addWidget(btn_uncheck_all)
        btn_add = QPushButton("  הוסף אדם")
        btn_add.setStyleSheet(_BTN_ACCENT)
        btn_add.setIcon(QIcon(line_icon("plus", 18, "#7c2d12")))
        btn_add.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_add.setToolTip("מוסיף לרשימה מקבל שאינו בה השבוע")
        btn_add.clicked.connect(self._add_person)
        tools.addWidget(btn_add)
        lf_lay.addLayout(tools)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["", "שם", "טלפון", "סטטוס"])
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        # Fixed width — several numbers can share one cell (#gaira).
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(2, 340)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.table.verticalHeader().setDefaultSectionSize(40)
        # The table grows to show ALL rows — the page scrolls, not a tiny
        # inner window (see _fit_table_height).
        self.table.setMinimumHeight(120)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        self.table.itemChanged.connect(self._on_item_changed)
        lf_lay.addWidget(self.table)
        hint = QLabel("☑ = יקבל צינתוק · כל המספרים של כל משפחה מצולצלים · "
                      "מספר משותף לשני מקבלים מצולצל פעם אחת · חריגים לא נשלחים")
        hint.setStyleSheet("color:#94a3b8; font-size:12px; " + _LBL)
        hint.setWordWrap(True)
        lf_lay.addWidget(hint)
        self.list_frame.setVisible(False)
        c_lay.addWidget(self.list_frame)
        lay.addWidget(card)

        # ── ② ההודעה ─────────────────────────────────────────────────────────
        card, c_lay, c_head = _step_card(
            "2", "ההודעה המושמעת", "מה ישמעו בטלפון")
        btn_test = QPushButton("  שלח בדיקה למספר שלי")
        btn_test.setStyleSheet(_BTN_GHOST)
        btn_test.setIcon(QIcon(line_icon("phone", 18, "#475569")))
        btn_test.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_test.setToolTip("מצלצל רק אליך, כדי לשמוע איך ההודעה נשמעת לפני "
                            "השליחה לכולם")
        btn_test.clicked.connect(self._send_test)
        c_head.addWidget(btn_test)

        self.lbl_rec = QLabel("")
        self.lbl_rec.setStyleSheet("color:#334155; font-size:13px; " + _LBL)
        self.lbl_rec.setWordWrap(True)
        c_lay.addWidget(self.lbl_rec)
        rec_row = QHBoxLayout()
        rec_row.setSpacing(8)
        btn_tts = QPushButton("🎙  צור הקלטה מטקסט…")
        btn_tts.setStyleSheet(_BTN_GHOST)
        btn_tts.setToolTip("כותבים את ההודעה — והמחשב מקריא אותה בקול טבעי "
                           "(חינם). ההקלטה נשמרת במאגר ומועלית לצינתוק.")
        btn_tts.clicked.connect(self._create_from_text)
        rec_row.addWidget(btn_tts)
        btn_upload = QPushButton("  העלה קובץ הקלטה…")
        btn_upload.setStyleSheet(_BTN_GHOST)
        btn_upload.setIcon(QIcon(line_icon("upload", 18, "#475569")))
        btn_upload.setToolTip("קובץ שמע (WAV/MP3) שיושמע בצינתוק — מומר אוטומטית "
                              "לפורמט הטלפוני בשרת של ימות")
        btn_upload.clicked.connect(self._upload_recording)
        rec_row.addWidget(btn_upload)
        self.btn_library = QPushButton("🎵  מאגר הקלטות")
        self.btn_library.setStyleSheet(_BTN_GHOST)
        self.btn_library.setToolTip("הקלטות קודמות ששמורות במחשב — אפשר להשמיע "
                                    "או להעלות שוב בלי להקליט מחדש")
        self.btn_library.clicked.connect(self._open_library)
        rec_row.addWidget(self.btn_library)
        rec_row.addStretch()
        btn_publish = QPushButton("פרסם בשלוחה 1…")
        btn_publish.setStyleSheet(_BTN_LINK)
        btn_publish.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_publish.setToolTip("מעתיק את ההודעה הנוכחית לשלוחת ההודעות של הקו — "
                               "שם כל מתקשר ישמע אותה. לא קורה אוטומטית אף פעם; "
                               "רק בלחיצה כאן ואחרי אישור.")
        btn_publish.clicked.connect(self._publish_to_line)
        rec_row.addWidget(btn_publish)
        c_lay.addLayout(rec_row)
        tip = QLabel("💡 כדאי לומר בסוף ההקלטה: \"לאישור הגעה חייגו חזרה לקו "
                     f"והקישו {yemot.SURVEY_EXT}: 1 מגיע, 2 לא מגיע, 3 לא יודע. "
                     "מי שלא יקיש — ייחשב כמי שלא שיתף פעולה\". התשובות מופיעות "
                     "בתוכנה ליד כל שם.")
        tip.setStyleSheet("color:#94a3b8; font-size:12px; " + _LBL)
        tip.setWordWrap(True)
        c_lay.addWidget(tip)
        lay.addWidget(card)

        # ── היסטוריה ─────────────────────────────────────────────────────────
        card, c_lay, c_head = _step_card(
            "", "היסטוריית צינתוקים", "כל השליחות, משני המחשבים")
        btn_hist_xls = QPushButton("  ייצוא לאקסל")
        btn_hist_xls.setStyleSheet(_BTN_GHOST)
        btn_hist_xls.setIcon(QIcon(line_icon("export", 18, "#475569")))
        btn_hist_xls.setToolTip("שומר קובץ אקסל עם כל הצינתוקים: סיכום לכל שליחה "
                                "+ פירוט לכל מספר — מי קיבל, מה ענה בסקר "
                                "(מגיע / לא מגיע / לא יודע / לא הגיב) ומי לא נענה")
        btn_hist_xls.clicked.connect(self._export_history)
        btn_ans = QPushButton("  רענן תשובות")
        btn_ans.setStyleSheet(_BTN_GHOST)
        btn_ans.setIcon(QIcon(line_icon("refresh", 18, "#475569")))
        btn_ans.setToolTip("קורא מהקו את התשובות שהוקשו בסקר (שלוחה "
                           f"{yemot.SURVEY_EXT}: 1 מגיע / 2 לא מגיע / 3 לא יודע) "
                           "לצינתוקים של השבועיים האחרונים ומעדכן את הטבלה, "
                           "ההיסטוריה ורשימת החלוקה")
        btn_ans.clicked.connect(self._refresh_answers)
        btn_srv = QPushButton("  היסטוריה מהשרת")
        btn_srv.setStyleSheet(_BTN_GHOST)
        btn_srv.setIcon(QIcon(line_icon("download", 18, "#475569")))
        btn_srv.setToolTip("מושך מהשרת של ימות את כל הקמפיינים שהקו הריץ אי-פעם "
                           "(גם לפני התוכנה) ואת יומן השיחות הנכנסות לקו — "
                           "כדי שהתוכנה תדע באיזו שעה כל אחד באמת עונה או מתקשר. "
                           "רץ לבד פעם ביום ברקע; הכפתור מרענן עכשיו.")
        btn_srv.clicked.connect(lambda: self._sync_history(manual=True))
        self.btn_hist_sync = btn_srv
        c_head.addWidget(btn_srv)
        c_head.addWidget(btn_ans)
        c_head.addWidget(btn_hist_xls)
        self.lbl_hist_sync = QLabel("")
        self.lbl_hist_sync.setStyleSheet(_LBL)
        self.lbl_hist_sync.setWordWrap(True)
        c_lay.addWidget(self.lbl_hist_sync)
        self._hist_worker = None
        self._hist_prog = ""
        self._refresh_hist_sync_label()
        self.hist = QTableWidget(0, 6)
        self.hist.setHorizontalHeaderLabels(
            ["מתי", "שם", "נשלחו", "הצליחו", "תשובות בסקר", "נכשלו"])
        self.hist.verticalHeader().setVisible(False)
        self.hist.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.hist.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        hh2 = self.hist.horizontalHeader()
        hh2.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh2.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for c in (2, 3, 4, 5):
            hh2.setSectionResizeMode(c, QHeaderView.ResizeMode.ResizeToContents)
        self.hist.verticalHeader().setDefaultSectionSize(34)
        self.hist.setMinimumHeight(150)
        self.hist.setMaximumHeight(260)
        c_lay.addWidget(self.hist)
        lay.addWidget(card)
        lay.addStretch()

        # ── Sticky bottom bar: ③ שליחה ───────────────────────────────────────
        bottom_wrap = QWidget()
        bw = QVBoxLayout(bottom_wrap)
        bw.setContentsMargins(20, 4, 20, 12)
        bw.setSpacing(0)
        bottom_bar = QFrame()
        bottom_bar.setObjectName("bottom-bar")
        bottom_bar.setStyleSheet(
            "QFrame#bottom-bar{background:#ffffff; border:1px solid #e6eaf2;"
            " border-radius:14px;}")
        bar = QVBoxLayout(bottom_bar)
        bar.setContentsMargins(16, 8, 16, 8)
        bar.setSpacing(6)

        # Scheduled-campaign strip (#xi85i) — visible while a schedule waits.
        self.sched_frame = QFrame()
        self.sched_frame.setStyleSheet(
            "QFrame{background:#fdf7e7; border:1px solid #efdead; border-radius:9px;}")
        sc_lay = QHBoxLayout(self.sched_frame)
        sc_lay.setContentsMargins(12, 5, 12, 5)
        self.lbl_sched = QLabel("")
        self.lbl_sched.setStyleSheet("color:#8a6410; font-weight:600; " + _LBL)
        self.lbl_sched.setWordWrap(True)
        sc_lay.addWidget(self.lbl_sched, 1)
        self.btn_cancel_sched = QPushButton("בטל תזמון")
        self.btn_cancel_sched.setStyleSheet(_BTN_GHOST)
        self.btn_cancel_sched.clicked.connect(self._cancel_sched)
        sc_lay.addWidget(self.btn_cancel_sched)
        self.sched_frame.setVisible(False)
        bar.addWidget(self.sched_frame)

        # Live-progress strip (hidden until a campaign runs).
        self.prog_frame = QFrame()
        self.prog_frame.setStyleSheet(
            "QFrame{background:#f0f9f6; border:1px solid #cfe9df; border-radius:9px;}")
        p_lay = QVBoxLayout(self.prog_frame)
        p_lay.setContentsMargins(12, 6, 12, 6)
        p_lay.setSpacing(4)
        self.lbl_prog = QLabel("")
        self.lbl_prog.setStyleSheet("font-weight:600; color:#0f766e; " + _LBL)
        self.lbl_prog.setWordWrap(True)
        p_lay.addWidget(self.lbl_prog)
        self.progress = QProgressBar()
        self.progress.setTextVisible(False)
        self.progress.setFixedHeight(10)
        p_lay.addWidget(self.progress)
        counters = QHBoxLayout()
        counters.setSpacing(14)
        self.lbl_ans = QLabel("")          # v3.02 — survey answers, rich text
        self.lbl_ans.setStyleSheet("font-weight:700; " + _LBL)
        self.lbl_ans.setTextFormat(Qt.TextFormat.RichText)
        self.lbl_done = QLabel("")
        self.lbl_done.setStyleSheet("color:#0f6e56; font-weight:600; " + _LBL)
        self.lbl_fail = QLabel("")
        self.lbl_fail.setStyleSheet("color:#a32d2d; font-weight:600; " + _LBL)
        self.lbl_wait = QLabel("")
        self.lbl_wait.setStyleSheet("color:#5f6d69; " + _LBL)
        for w in (self.lbl_ans, self.lbl_done, self.lbl_fail, self.lbl_wait):
            counters.addWidget(w)
        counters.addStretch()
        self.btn_extend_track = QPushButton("⏱ הארך מעקב ב-30 דק'")
        self.btn_extend_track.setStyleSheet(_BTN_GHOST)
        self.btn_extend_track.clicked.connect(self._extend_tracking)
        self.btn_extend_track.setVisible(False)
        counters.addWidget(self.btn_extend_track)
        self.btn_stop_track = QPushButton("סיים מעקב")
        self.btn_stop_track.setStyleSheet(_BTN_GHOST)
        self.btn_stop_track.setToolTip("סוגר את חלון המעקב עכשיו ושומר את מה "
                                       "שנאסף עד כה בהיסטוריה")
        self.btn_stop_track.clicked.connect(self._stop_tracking_now)
        self.btn_stop_track.setVisible(False)
        counters.addWidget(self.btn_stop_track)
        self.btn_resend = QPushButton("🔄 שלח שוב לנכשלים")
        self.btn_resend.setStyleSheet(_BTN_ACCENT)
        self.btn_resend.clicked.connect(self._resend_failed)
        self.btn_resend.setVisible(False)
        counters.addWidget(self.btn_resend)
        p_lay.addLayout(counters)
        self.prog_frame.setVisible(False)
        bar.addWidget(self.prog_frame)

        act = QHBoxLayout()
        act.setSpacing(12)
        act.addWidget(_step_badge("3"))
        self.lbl_summary = QLabel("")
        self.lbl_summary.setStyleSheet("color:#334155; font-size:14px; font-weight:700; " + _LBL)
        act.addWidget(self.lbl_summary)
        act.addStretch()
        self.btn_sched = QPushButton("  תזמן שליחה…")
        self.btn_sched.setStyleSheet(_BTN_GHOST)
        self.btn_sched.setIcon(QIcon(line_icon("calendar", 18, "#475569")))
        self.btn_sched.setMinimumHeight(46)
        self.btn_sched.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_sched.setToolTip("קובעים תאריך ושעה — והצינתוק יוצא לבד "
                                  "מהשרת של ימות המשיח, גם כשהמחשב כבוי")
        self.btn_sched.clicked.connect(self._schedule)
        act.addWidget(self.btn_sched)
        self.btn_send = QPushButton("")
        self.btn_send.setObjectName("primary")
        self.btn_send.setStyleSheet(_BTN_PRINT)
        self.btn_send.setIcon(QIcon(line_icon("send", 20, "#ffffff")))
        self.btn_send.setMinimumWidth(240)
        self.btn_send.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_send.clicked.connect(self._send)
        act.addWidget(self.btn_send)
        bar.addLayout(act)
        bw.addWidget(bottom_bar)
        s_lay.addWidget(bottom_wrap, 0)

    # ── List building ─────────────────────────────────────────────────────────

    def refresh(self):
        """Rebuild the call list from the CURRENT distribution list (the same
        rows the 'חלוקה ורישום' screen shows, minus the reserve section) — or,
        when a past distribution was loaded (#9hgvi), from that batch."""
        configured = yemot.is_configured()
        self.banner.setVisible(not configured)
        self.lbl_ok.setVisible(configured)
        self._refresh_batch_banner()
        # #ifc70 — nothing is loaded until the operator asks for a list.
        loaded = (self._batch is not None or self._free is not None
                  or self._list_loaded)
        self.load_frame.setVisible(not loaded)
        self.list_frame.setVisible(loaded)
        base = self._distribution_rows() if loaded else []
        manual = [r for r in self._rows if r.get("manual")]
        manual_ids = {r["rec"].get("id") for r in manual}
        self._rows = []
        for rec in base:
            if rec.get("id") in manual_ids:
                manual = [m for m in manual if m["rec"].get("id") != rec.get("id")]
            self._rows.append(self._make_row(rec))
        self._rows.extend(manual)      # keep hand-added people across refreshes
        try:                           # #y7jr0 — tooltips of answer history
            self._stats = yemot.answer_stats()
        except Exception:
            self._stats = {}
        self._flag_duplicates()
        self._populate()
        self._refresh_recording_label()
        self._refresh_history()
        self._maybe_resume_tracking()

    def _distribution_rows(self):
        if self._free is not None:             # standalone list (#1/9)
            return [{"id": None, "full_name": name or "", "phone1": phone}
                    for phone, name in self._free]
        if self._batch is not None:            # past distribution (#9hgvi)
            try:
                return db.get_batch_export_rows(self._batch.get("id"))
            except Exception:
                return []
        gt = getattr(self.main, "group_tab", None)
        if gt is None:
            return []
        try:
            if not gt._rows_data:
                gt.refresh()
        except Exception:
            pass
        reserve_ids = getattr(gt, "_reserve_ids", set()) or set()
        return [dict(r) for r in (gt._rows_data or [])
                if not r.get("_reserve") and r.get("id") not in reserve_ids]

    # ── Past-distribution mode (#9hgvi) ───────────────────────────────────────

    def load_batch(self, batch: dict):
        """Show a PAST distribution's recipients as the call list — entry point
        of the right-click 'שלח צינתוק' action in 'חלוקות קודמות'."""
        self._batch = dict(batch or {})
        self._free = None
        self._rows = []                # drop week-list rows and manual picks
        self._last_entries = []
        self.refresh()

    def _load_week_list(self):
        """#ifc70 — explicit load of the current distribution list."""
        self._list_loaded = True
        self._free = None
        self.refresh()

    def _load_free_list(self):
        """#1/9 — a standalone pasted/Excel list, unrelated to distributions."""
        dlg = _FreeListDialog(self)
        if not dlg.exec() or not dlg.entries:
            return
        self._free = list(dlg.entries)
        self._batch = None
        self._rows = []
        self._last_entries = []
        self.refresh()

    def _clear_batch(self):
        self._batch = None
        self._free = None
        self._rows = []
        self._last_entries = []
        self._list_loaded = False    # back to the explicit-load state (#ifc70)
        self.refresh()

    def _refresh_batch_banner(self):
        if self._free is not None:
            self.lbl_batch.setText(f"📋 רשימה עצמאית — {len(self._free)} "
                                   "מספרים (בלי קשר לרשימות החלוקה)")
            self.btn_back.setText("נקה את הרשימה")
            self.batch_frame.setVisible(True)
            return
        self.btn_back.setText("חזור לרשימת השבוע")
        if self._batch is None:
            self.batch_frame.setVisible(False)
            return
        name = self._batch.get("dist_name") or ""
        date = self._batch.get("dist_date") or ""
        if date and len(date) >= 10 and date[4] == "-":
            date = f"{date[8:10]}/{date[5:7]}/{date[:4]}"
        self.lbl_batch.setText(f"📋 הרשימה נטענה מחלוקה קודמת: {name or date}"
                               + (f" ({date})" if name and date else ""))
        self.batch_frame.setVisible(True)

    def _dist_date_iso(self) -> str:
        """The distribution date the campaign belongs to (double-send guard).
        A standalone list is not tied to any distribution — it gets an empty
        date, so the guard never blocks it (and it never blocks the week)."""
        if self._free is not None:
            return ""
        if self._batch is not None and (self._batch.get("dist_date") or ""):
            return self._batch["dist_date"]
        return db.next_wednesday().isoformat()

    def _campaign_name(self) -> str:
        if self._free is not None:
            return f"רשימה עצמאית — {len(self._free)} מספרים"
        if self._batch is not None:
            label = self._batch.get("dist_name") or self._dist_date_iso()
            return f"צינתוק לחלוקה — {label}"
        return f"חלוקה של {db.next_wednesday().strftime('%d/%m/%Y')}"

    # ── Row model ─────────────────────────────────────────────────────────────

    def _make_row(self, rec: dict, manual: bool = False) -> dict:
        phones = yemot.pick_phones(rec)
        raw_any = any((rec.get(f) or "").strip()
                      for f in ("phone1", "phone2", "phone3"))
        why = "" if phones else ("מספר לא תקין" if raw_any else "אין מספר טלפון")
        return {"rec": rec, "phones": phones, "send": list(phones),
                "checked": bool(phones), "why": why, "manual": manual}

    def _flag_duplicates(self):
        """All of a recipient's numbers ring (#gaira); a number shared by two
        rows rings only for the FIRST (one household — one call). A row left
        with no numbers at all becomes an unchecked exception."""
        allocated = yemot.allocate_phones([r["phones"] for r in self._rows])
        first_owner = {}
        for row, mine in zip(self._rows, allocated):
            for p in mine:
                first_owner.setdefault(p, row["rec"].get("full_name") or "")
            row["send"] = mine
            if row["phones"] and not mine:
                row["checked"] = False
                owner = first_owner.get(row["phones"][0], "")
                row["why"] = f"אותו מספר כמו {owner}"
            elif row["why"].startswith("אותו מספר"):
                row["why"] = ""

    def _phone_tooltip(self, row) -> str:
        """#y7jr0 שלב 1 — היסטוריית מענה לכל מספר של השורה, מתוך הדוחות
        השמורים; שעה מומלצת מופיעה אחרי MIN_SMART_HISTORY שליחות."""
        parts = []
        for p in row["phones"]:
            s = self._stats.get(p)
            if not s:
                continue
            line = f"{p}: ענה ב-{s['answered']} מתוך {s['attempts']} צינתוקים"
            if s.get("calls"):
                line += f" · התקשר חזרה לקו {s['calls']} פעמים"
                usual = yemot.usual_call_hour(s)
                if usual is not None:
                    line += f" (בדרך כלל בסביבות {usual:02d}:00)"
            if s.get("best_hour") is not None:
                line += f" · השעה הכי טובה להשיג אותו: {s['best_hour']:02d}:00"
            else:
                line += (f" (שעה מומלצת תוצג אחרי "
                         f"{yemot.MIN_SMART_HISTORY} שליחות או "
                         f"{yemot.MIN_CALLBACK_HISTORY} התקשרויות חוזרות)")
            parts.append(line)
        return "\n".join(parts)

    def _populate(self):
        self.table.blockSignals(True)
        # Full reset: clearContents also drops leftover cell widgets from the
        # pre-v2.88 phone combos (they would float over the repopulated table).
        self.table.clearContents()
        self.table.setRowCount(len(self._rows))
        for i, row in enumerate(self._rows):
            rec = row["rec"]
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            chk.setCheckState(Qt.CheckState.Checked if row["checked"]
                              else Qt.CheckState.Unchecked)
            if not row["phones"]:
                chk.setFlags(Qt.ItemFlag.NoItemFlags)
            self.table.setItem(i, 0, chk)
            name = QTableWidgetItem(rec.get("full_name") or "")
            if row["manual"]:
                name.setText((rec.get("full_name") or "") + "  (נוסף ידנית)")
            self.table.setItem(i, 1, name)
            shown = row["send"] or row["phones"]
            phones_item = QTableWidgetItem(", ".join(shown) if shown else "—")
            tip = self._phone_tooltip(row)
            if tip:
                phones_item.setToolTip(tip)
                name.setToolTip(tip)
            self.table.setItem(i, 2, phones_item)
            status = QTableWidgetItem("מוכן" if row["checked"] and row["send"]
                                      else ("⚠ " + row["why"] if row["why"] else "לא נשלח"))
            if row["why"]:
                status.setForeground(Qt.GlobalColor.darkYellow)
            self.table.setItem(i, 3, status)
        self.table.blockSignals(False)
        if self._last_entries:
            # A refresh (tab switch / sync) rebuilt the rows — restore the
            # per-row results of the campaign still being tracked / just done.
            self._apply_results_to_table(self._last_entries)
        self._fit_table_height()
        self._update_metrics()

    def _fit_table_height(self):
        """#1/9: size the table to show ALL rows — the page scrolls instead of
        a tiny inner window (Ron saw only ~4 families at a time)."""
        h = self.table.horizontalHeader().height() + 6
        for r in range(self.table.rowCount()):
            h += self.table.rowHeight(r)
        self.table.setMinimumHeight(max(120, h))

    def _set_all_checked(self, state: bool):
        """#1/9: one click instead of un/checking V row by row. Checking marks
        only rows that actually have a number to ring."""
        self.table.blockSignals(True)
        for i, row in enumerate(self._rows):
            row["checked"] = bool(state and row["send"])
            it = self.table.item(i, 0)
            if it is not None and it.flags() & Qt.ItemFlag.ItemIsUserCheckable:
                it.setCheckState(Qt.CheckState.Checked if row["checked"]
                                 else Qt.CheckState.Unchecked)
        self.table.blockSignals(False)
        self._update_metrics()

    def _on_item_changed(self, item):
        if item.column() != 0:
            return
        i = item.row()
        if 0 <= i < len(self._rows):
            self._rows[i]["checked"] = item.checkState() == Qt.CheckState.Checked
            self._update_metrics()

    def _update_metrics(self):
        total = len(self._rows)
        ready = len(self._ready_rows())
        bad = sum(1 for r in self._rows if r["why"])
        _set_metric(self.m_total, total)
        _set_metric(self.m_ready, ready)
        _set_metric(self.m_bad, bad)
        busy = self._worker is not None or self._cb_worker is not None
        if not self._rows:
            self.lbl_summary.setText("שליחה — טען קודם רשימת נמענים")
        elif ready:
            self.lbl_summary.setText(f"שליחה — {ready} משפחות מסומנות"
                                     + (f", {bad} חריגים לא יישלחו" if bad else ""))
        else:
            self.lbl_summary.setText("שליחה — אף נמען לא מסומן")
        self.btn_send.setText(f"  שלח צינתוק ל-{ready}" if ready else "  שלח צינתוק")
        self.btn_send.setEnabled(ready > 0 and not busy)
        self.btn_sched.setEnabled(ready > 0 and not busy)

    def _ready_rows(self):
        return [r for r in self._rows if r["checked"] and r["send"]]

    def _phones_map(self, rows) -> dict:
        """{'0501234567': 'שם', …} — every number of every checked row (#gaira);
        cross-row duplicates were already removed by _flag_duplicates."""
        phones = {}
        for r in rows:
            for p in r["send"]:
                phones.setdefault(p, r["rec"].get("full_name") or "")
        return phones

    # ── Actions ───────────────────────────────────────────────────────────────

    def _run_blocking(self, fn, text="מתחבר לשרת ימות המשיח…"):
        """Run one server call off the UI thread behind a small modal "working"
        dialog, and return its result (or re-raise its exception) as if it ran
        inline. The window keeps repainting, so a slow/absent connection (up
        to ~2.5 minutes of retries) no longer shows Windows' "לא מגיב", and the
        operator cannot click anything else meanwhile."""
        dlg = QProgressDialog(text, "", 0, 0, self)
        dlg.setCancelButton(None)
        dlg.setWindowTitle("צינתוקים")
        dlg.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dlg.setWindowModality(Qt.WindowModality.ApplicationModal)
        dlg.setMinimumDuration(300)     # instant answers never flash a dialog
        dlg.setValue(0)
        box = {}
        loop = QEventLoop(self)
        worker = _TaskWorker(fn, self)
        worker.done.connect(lambda r: (box.__setitem__("r", r), loop.quit()))
        worker.finished.connect(loop.quit)
        worker.start()
        loop.exec()
        worker.wait(2000)
        worker.deleteLater()
        dlg.reset()
        dlg.deleteLater()
        res = box.get("r")
        if isinstance(res, Exception):
            raise res
        return res

    def _goto_settings(self):
        if self.main and hasattr(self.main, "navigate_to_tab"):
            self.main.navigate_to_tab(self.main.settings_tab)

    def _add_person(self):
        dlg = _AddPersonDialog({r["rec"].get("id") for r in self._rows}, self)
        if dlg.exec() and dlg.picked:
            self._rows.append(self._make_row(dlg.picked, manual=True))
            self._flag_duplicates()
            self._populate()

    def _refresh_recording_label(self):
        tid = (db.get_setting(yemot.SET_TEMPLATE) or "").strip()
        confirm_tip = ("\n💡 כדאי לומר בסוף ההקלטה: \"לאישור הגעה — הקש 7, "
                       "לשמיעה חוזרת — הקש 1\". מי שיקיש 7 יסומן בתוכנה "
                       "כ\"אישר הגעה\" (המקשים קבועים ע\"י ימות המשיח).")
        if not yemot.is_configured():
            self.lbl_rec.setText("ההודעה המושמעת: תוגדר אחרי חיבור המערכת (בהגדרות).")
        elif tid:
            self.lbl_rec.setText(
                f"ההודעה המושמעת שמורה בימות המשיח (תבנית {tid}). "
                "להחלפה — צור הקלטה מטקסט / העלה קובץ / בחר מהמאגר, "
                "ואז \"שלח בדיקה\" כדי לשמוע אותה." + confirm_tip)
        else:
            self.lbl_rec.setText(
                "עוד לא הוגדרה הודעה: צור הקלטה מטקסט (המחשב מקריא בקול טבעי) "
                "או העלה קובץ הקלטה." + confirm_tip)

    def _require_config(self) -> bool:
        if yemot.is_configured():
            return True
        QMessageBox.information(
            self, "צינתוקים",
            "עוד לא הוזנו פרטי הגישה לימות המשיח.\n"
            "בהגדרות ← \"צינתוקים (ימות המשיח)\" — הזן מספר מערכת וסיסמה.")
        self._goto_settings()
        return False

    def _upload_path(self, path: str, title: str) -> bool:
        """מעלה קובץ שמע לתבנית הקמפיין בימות; מציג הודעה. True = הצליח."""
        try:
            self._run_blocking(lambda: yemot.upload_message_wav(path), "מעלה את ההקלטה לשרת…")
            ok, msg = True, ("ההקלטה הועלתה בהצלחה ✓\n"
                             "שלח בדיקה למספר שלך כדי לשמוע אותה בטלפון.")
        except yemot.YemotError as e:
            ok, msg = False, str(e)
        except Exception as e:
            ok, msg = False, f"ההעלאה נכשלה: {e}"
        (QMessageBox.information if ok else QMessageBox.warning)(self, title, msg)
        self._refresh_recording_label()
        return ok

    def _upload_recording(self):
        if not self._require_config():
            return
        path, _f = QFileDialog.getOpenFileName(
            self, "בחר קובץ הקלטה", "", "קובצי שמע (*.wav *.mp3);;כל הקבצים (*.*)")
        if not path:
            return
        if self._upload_path(path, "העלאת הקלטה"):
            # נשמר גם במאגר (#kgmcw) — שלא יצטרכו לחפש את הקובץ שוב בפעם הבאה.
            try:
                name = os.path.splitext(os.path.basename(path))[0]
                tts.library_add(path, name, source="file")
            except OSError:
                pass

    def _create_from_text(self):
        """#9vy1b — יצירת הקלטה מטקסט בקול AI חינמי, שמירה במאגר והעלאה."""
        if not self._require_config():
            return
        dlg = TtsDialog(self)
        if not dlg.exec() or not dlg.result_path:
            return
        try:
            item = tts.library_add(dlg.result_path, tts.suggest_name(dlg.result_text),
                                   text=dlg.result_text, voice=dlg.result_voice,
                                   source="tts")
            path = tts.library_path(item)
        except OSError as e:
            QMessageBox.warning(self, "יצירת הקלטה", f"שמירת ההקלטה במאגר נכשלה: {e}")
            path = dlg.result_path
        if dlg.result_note:
            QMessageBox.information(self, "יצירת הקלטה", dlg.result_note)
        self._upload_path(path, "יצירת הקלטה")

    def _open_library(self):
        dlg = LibraryDialog(self)
        if dlg.exec() and dlg.picked_path:
            if not self._require_config():
                return
            self._upload_path(dlg.picked_path, "מאגר הקלטות")

    def _send_test(self):
        # #dx28e — the test number is entered/edited right here (pre-filled with
        # the saved one) and remembered; no settings-screen field any more.
        if not self._require_config():
            return
        saved = (db.get_setting(yemot.SET_TEST_PHONE) or "").strip()
        phone, okd = QInputDialog.getText(
            self, "שליחת בדיקה",
            "לאיזה מספר לצלצל? (המספר שלך — נשמר לפעם הבאה)", text=saved)
        if not okd:
            return
        phone = phone.strip()
        if not yemot.normalize_phone(phone):
            QMessageBox.warning(self, "שליחת בדיקה", "המספר שהוזן אינו תקין.")
            return
        db.set_setting(yemot.SET_TEST_PHONE, phone)
        try:
            res = self._run_blocking(lambda: yemot.run_test(phone), "שולח שיחת בדיקה…")
        except yemot.YemotError as e:
            QMessageBox.warning(self, "שליחת בדיקה", str(e))
            return
        except Exception as e:
            QMessageBox.warning(self, "שליחת בדיקה", f"השליחה נכשלה: {e}")
            return
        cid = str(res.get("campaignId") or "")
        if cid:
            # מעקב חי — מספר הבדיקה לא תמיד ביד המפעיל, אז מראים כאן אם
            # השיחה נענתה / נכשלה / הוקש 7.
            _TestStatusDialog(cid, phone, self).exec()
        else:
            QMessageBox.information(
                self, "שליחת בדיקה",
                f"📞 מצלצל עכשיו ({phone}) — מי שעונה ישמע את ההודעה.\n"
                "אפשר גם לא לענות ולהתקשר חזרה לקו — ההודעה תושמע בכניסה.")

    def _send(self):
        if not self._require_config():
            return
        rows = self._ready_rows()
        phones = self._phones_map(rows)
        if not phones:
            QMessageBox.warning(self, "צינתוקים", "אין אף נמען מסומן עם מספר תקין.")
            return
        # A pending schedule dials the list STORED in the template — and an
        # immediate send replaces that list. Refuse instead of silently
        # re-targeting the scheduled campaign at these people.
        pending = self._pending_sched()
        if pending:
            when = timefmt.datetime_str(pending.get("sent_at") or "")
            QMessageBox.information(
                self, "צינתוקים",
                f"קיים צינתוק מתוזמן ({when}).\n"
                "שליחה עכשיו הייתה מחליפה את רשימת הנמענים של התזמון — "
                "והצינתוק המתוזמן היה יוצא לאנשים האלה במקום לרשימה המקורית.\n\n"
                "בטל קודם את התזמון (כפתור \"בטל תזמון\"), שלח, "
                "ואז תזמן מחדש אם צריך.")
            return
        dist_date = self._dist_date_iso()
        prev = db.tzintuk_campaign_for_date(dist_date)
        extra = ""
        if prev:
            when = timefmt.datetime_str(prev.get("sent_at") or "")
            src = prev.get("device") or "מחשב אחר"
            verb = ("כבר מתוזמן צינתוק"
                    if prev.get("status") == "scheduled" else "כבר נשלח צינתוק")
            extra = (f"\n\n⚠ שים לב: {verb} לחלוקה של תאריך זה "
                     f"({when}, מ{src}). שליחה נוספת תצלצל לאנשים פעם שנייה!")
        bad = sum(1 for r in self._rows if r["why"])
        summary = (f"{self._campaign_name()}\n\n"
                   f"עומד לשלוח צינתוק ל-{len(rows)} משפחות "
                   f"({len(phones)} מספרי טלפון — כל המספרים של כל משפחה).\n"
                   f"חריגים שלא יישלחו: {bad}.\n"
                   "מי שלא ענה ישמע את ההודעה כשיתקשר חזרה לקו "
                   "(מושמעת רק למי שברשימה הזו).\n"
                   f"מי שיחייג חזרה ויקיש {yemot.SURVEY_EXT} יסומן בתוכנה לפי "
                   "תשובתו: 1 מגיע / 2 לא מגיע / 3 לא יודע; "
                   f"מי שלא יקיש — \"לא הגיב\".{extra}")
        dlg = _SendModeDialog(summary, self)
        if not dlg.exec() or not dlg.mode:
            return
        classic = dlg.mode == "classic"
        self.btn_send.setEnabled(False)      # locked until the campaign ends
        self.btn_send.setText("שולח…")
        def _do_send():
            tid = yemot.ensure_template()
            # store_list=True — הרשימה נשמרת בתבנית כדי שהקו ישמיע את
            # ההודעה למי שמתקשר חזרה (#z4xy9, campaign_message_to_play).
            return tid, yemot.run_campaign(phones, tid, store_list=True,
                                           classic=classic)
        try:
            template_id, res = self._run_blocking(
                _do_send, f"שולח צינתוק ל-{len(phones)} מספרים…")
        except yemot.YemotError as e:
            QMessageBox.warning(self, "צינתוקים", str(e))
            self._update_metrics()
            return
        except Exception as e:
            QMessageBox.warning(self, "צינתוקים", f"השליחה נכשלה: {e}")
            self._update_metrics()
            return
        from utils import sync
        name = self._campaign_name()
        sent_iso = datetime.now(timezone.utc).isoformat()   # survey answers count from now
        if classic:
            # v2.96 — קלאסי: אין מענה לצלצול עצמו, אבל עוקבים חי אחרי מי
            # שמתקשר חזרה לקו; התשובה בסקר (77) נקראת מקובץ הנתונים של
            # השלוחה. לימות אין יומן שיחות עבר, לכן מעקב-החזרה חי בלבד.
            tracker = yemot.CallbackTracker(phones)
            guid = db.add_tzintuk_campaign(
                "צינתוק קלאסי — " + name, dist_date, template_id,
                str(res.get("campaignId") or ""),
                int(res.get("entriesCount") or len(phones)),
                device=sync.device_name() or "", status="sending")
            db.update_tzintuk_campaign(
                guid, 0, 0, "sending",
                json.dumps(tracker.entries(), ensure_ascii=False))
            self._active_guid = guid
            self._refresh_history()
            if res.get("classic_fallback"):
                ring_line = ("⚠ שירות הצינתוק אינו פעיל בקו, לכן נשלח צלצול קצר "
                             f"({yemot.CLASSIC_RING_SECONDS} שניות) — מי שמספיק לענות "
                             "ישמע את ההודעה.\n")
            else:
                ring_line = "הטלפונים יצלצלו ויתנתקו — אי אפשר לענות לצלצול.\n"
            QMessageBox.information(
                self, "צינתוק קלאסי",
                f"📞 הצלצולים יצאו ל-{len(phones)} מספרים.\n" + ring_line +
                "מי שמתקשר חזרה לקו ישמע את ההודעה — ובחצי השעה הקרובה "
                f"תראה כאן בזמן אמת מי חזר לשיחה ומה ענה בסקר (הקשה {yemot.SURVEY_EXT}).")
            self._start_callback_tracking(
                guid, phones, time.time() + yemot.CLASSIC_TRACK_SECONDS,
                since_iso=sent_iso)
            return
        self._active_guid = db.add_tzintuk_campaign(
            name, dist_date, template_id,
            str(res.get("campaignId") or ""),
            int(res.get("entriesCount") or len(phones)),
            device=sync.device_name() or "")
        self._refresh_history()
        self._start_tracking(str(res.get("campaignId") or ""), len(phones), sent_iso)

    def _publish_to_line(self):
        """#kx6wd — copy the current campaign recording into the line's central
        messages extension (שלוחה 1). MANUAL ONLY (user decision 31/08/2026):
        the extension is heard by EVERY caller, and a tzintuk to aid recipients
        is private — so nothing is ever published automatically; this runs only
        from the explicit button, behind a clear warning."""
        if not self._require_config():
            return
        ans = QMessageBox.question(
            self, "פרסום בשלוחת ההודעות",
            "⚠ שים לב: שלוחה 1 פתוחה לכל מי שמתקשר לקו.\n"
            "ההודעה הנוכחית של הצינתוק תתפרסם שם כהודעה נוספת, "
            "וכל שומעי הקו יוכלו לשמוע אותה — לא רק מקבלי החלוקה.\n\n"
            "לפרסם הודעה זו לכולם?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if ans != QMessageBox.StandardButton.Yes:
            return
        try:
            name = self._run_blocking(yemot.publish_to_extension, "מפרסם בשלוחה 1…")
            ok, msg = True, ("ההודעה פורסמה בשלוחה 1 ✓ "
                             f"(קובץ {name}).")
        except Exception as e:
            ok, msg = False, f"הפרסום נכשל:\n{e}"
        (QMessageBox.information if ok else QMessageBox.warning)(
            self, "פרסום בשלוחת ההודעות", msg)

    def _resend_failed(self):
        if not self._last_failed:
            return
        phones = {e["phone"]: e.get("name") or "" for e in self._last_failed
                  if e.get("phone")}
        if not phones:
            return
        ans = QMessageBox.question(
            self, "שליחה חוזרת",
            f"לשלוח שוב ל-{len(phones)} שנכשלו בסבב הקודם?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans != QMessageBox.StandardButton.Yes:
            return
        self.btn_resend.setVisible(False)
        try:
            template_id, res = self._run_blocking(
                lambda: (lambda t: (t, yemot.run_campaign(phones, t)))(yemot.ensure_template()),
                f"שולח שוב ל-{len(phones)} מספרים…")
        except Exception as e:
            QMessageBox.warning(self, "צינתוקים", f"השליחה נכשלה: {e}")
            return
        from utils import sync
        dist_date = self._dist_date_iso()
        self._active_guid = db.add_tzintuk_campaign(
            f"שליחה חוזרת לנכשלים ({len(phones)})", dist_date, template_id,
            str(res.get("campaignId") or ""), len(phones),
            device=sync.device_name() or "")
        self._refresh_history()
        self._start_tracking(str(res.get("campaignId") or ""), len(phones))

    # ── Scheduled campaigns (#xi85i) ─────────────────────────────────────────

    @staticmethod
    def _to_utc_iso(when: datetime) -> str:
        """Naive Israel-local time → UTC iso stamp (the sent_at convention)."""
        zone = timefmt._israel_zone()
        aware = when.replace(tzinfo=zone) if zone is not None else when.astimezone()
        return aware.astimezone(timezone.utc).isoformat()

    @staticmethod
    def _sched_dt(camp: dict):
        """The planned run time of a scheduled record, as aware Israel time."""
        return timefmt.to_israel(camp.get("sent_at") or "")

    def _pending_sched(self):
        for c in db.get_tzintuk_campaigns(limit=25):
            if c.get("status") == "scheduled":
                return c
        return None

    def _smart_hint(self, phones) -> str:
        """#y7jr0 שלב 1 — ההמלצה בדיאלוג התזמון: השעה עם אחוז המענה הגבוה
        ביותר על פני כל המספרים שברשימה (רק ממי שכבר יש עליו היסטוריה)."""
        by_hour = {}
        for p in phones:
            s = self._stats.get(p)
            if not s:
                continue
            for h, (a, t) in s["by_hour"].items():
                bucket = by_hour.setdefault(h, [0, 0])
                bucket[0] += a
                bucket[1] += t
        best = [(a / t, t, h) for h, (a, t) in by_hour.items() if t >= 10]
        if not best:
            return ""
        rate, total, hour = max(best)
        calls = sum((self._stats.get(p) or {}).get("calls", 0) for p in phones)
        extra = f", כולל {calls} התקשרויות חוזרות לקו" if calls else ""
        return (f"לפי ההיסטוריה של הרשימה הזו, השעה שבה הכי קל להשיג אנשים היא "
                f"{hour:02d}:00 ({rate:.0%} הצלחה, {total} שיחות שנבדקו{extra}).")

    def _schedule(self):
        if not self._require_config():
            return
        rows = self._ready_rows()
        phones = self._phones_map(rows)
        if not phones:
            QMessageBox.warning(self, "תזמון שליחה",
                                "אין אף נמען מסומן עם מספר תקין.")
            return
        pending = self._pending_sched()
        if pending:
            when = timefmt.datetime_str(pending.get("sent_at") or "")
            QMessageBox.information(
                self, "תזמון שליחה",
                f"כבר קיים צינתוק מתוזמן ({when}).\n"
                "אפשר לתזמן רק צינתוק אחד בכל פעם — בטל אותו קודם "
                "(כפתור \"בטל תזמון\") ואז תזמן מחדש.")
            return
        dlg = _ScheduleDialog(len(phones), self, smart_hint=self._smart_hint(phones))
        if not dlg.exec() or dlg.when is None:
            return
        when = dlg.when
        dist_date = self._dist_date_iso()
        prev = db.tzintuk_campaign_for_date(dist_date)
        extra = ""
        if prev:
            prev_when = timefmt.datetime_str(prev.get("sent_at") or "")
            verb = ("כבר מתוזמן צינתוק"
                    if prev.get("status") == "scheduled" else "כבר נשלח צינתוק")
            extra = (f"\n\n⚠ שים לב: {verb} לחלוקה של תאריך זה ({prev_when}). "
                     "שליחה נוספת תצלצל לאנשים פעם שנייה!")
        bad = sum(1 for r in self._rows if r["why"])
        ans = QMessageBox.question(
            self, "אישור תזמון",
            f"הצינתוק יישלח ל-{len(phones)} נמענים "
            f"ביום {when.strftime('%d/%m/%Y')} בשעה {when.strftime('%H:%M')}.\n"
            f"חריגים שלא יישלחו: {bad}.\n"
            f"עלות משוערת: כ-{len(phones)} יחידות.\n"
            "השליחה תצא מהשרת של ימות המשיח גם אם המחשב יהיה כבוי; "
            f"התוצאות ייקלטו בתוכנה בהפעלה הבאה.{extra}\n\nלתזמן?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if ans != QMessageBox.StandardButton.Yes:
            return
        try:
            template_id, res = self._run_blocking(
                lambda: (lambda t: (t, yemot.schedule_campaign(when, phones, t)))(yemot.ensure_template()),
                "שומר את התזמון בשרת…")
        except yemot.YemotError as e:
            QMessageBox.warning(self, "תזמון שליחה", str(e))
            return
        except Exception as e:
            QMessageBox.warning(self, "תזמון שליחה", f"התזמון נכשל: {e}")
            return
        from utils import sync
        db.add_tzintuk_campaign(
            f"צינתוק מתוזמן — {self._campaign_name()}",
            dist_date, template_id, str(res.get("schedId") or ""),
            int(res.get("count") or len(phones)),
            sent_at=self._to_utc_iso(when),
            device=sync.device_name() or "", status="scheduled")
        self._refresh_history()
        QMessageBox.information(
            self, "תזמון שליחה",
            f"נקבע ✓ — הצינתוק יישלח ביום {when.strftime('%d/%m/%Y')} "
            f"בשעה {when.strftime('%H:%M')}, גם אם המחשב יהיה כבוי.")

    def _refresh_sched_banner(self):
        camp = self._pending_sched()
        if camp is None:
            self.sched_frame.setVisible(False)
            return
        when = timefmt.datetime_str(camp.get("sent_at") or "")
        src = camp.get("device") or ""
        self.lbl_sched.setText(
            f"🕒 צינתוק מתוזמן ל-{when} — ל-{camp.get('total') or 0} נמענים"
            + (f" (נקבע מ{src})" if src else "")
            + ". המחשב לא חייב להיות דלוק בשעת השליחה.")
        self.sched_frame.setVisible(True)

    def _cancel_sched(self):
        camp = self._pending_sched()
        if camp is None:
            self.sched_frame.setVisible(False)
            return
        when = timefmt.datetime_str(camp.get("sent_at") or "")
        ans = QMessageBox.question(
            self, "ביטול תזמון",
            f"לבטל את הצינתוק המתוזמן ל-{when}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans != QMessageBox.StandardButton.Yes:
            return
        sched_id = camp.get("campaign_id") or ""
        try:
            self._run_blocking(lambda: yemot.delete_scheduled_campaign(sched_id),
                               "מבטל את התזמון בשרת…")
            db.update_tzintuk_campaign(camp["guid"], 0, 0, "canceled")
            ok, msg = True, "התזמון בוטל — לא יישלח צינתוק."
        except yemot.YemotError as e:
            if e.code == 106:      # already ran — go pick up the results
                ok, msg = False, str(e)
            elif e.code == 105:    # not on the server → nothing will ring
                db.update_tzintuk_campaign(camp["guid"], 0, 0, "canceled")
                ok, msg = True, "התזמון כבר לא קיים בשרת — סומן כמבוטל."
            else:
                ok, msg = False, str(e)
        except Exception as e:
            ok, msg = False, f"הביטול נכשל: {e}"
        (QMessageBox.information if ok else QMessageBox.warning)(
            self, "ביטול תזמון", msg)
        self._refresh_history()
        self._maybe_resume_tracking()

    def _check_scheduled(self):
        """A scheduled campaign whose time passed while the app was off (or
        closed): ask the server what happened, off the UI thread, and continue
        with the normal live tracking once the real campaignId is known."""
        self._refresh_sched_banner()
        if (self._sched_checker is not None or self._worker is not None
                or not yemot.is_configured()):
            return
        zone = timefmt._israel_zone()
        now = datetime.now(zone) if zone is not None else datetime.now().astimezone()
        due = []
        for c in db.get_tzintuk_campaigns(limit=25):
            if c.get("status") != "scheduled":
                continue
            dt = self._sched_dt(c)
            if dt is not None and now >= dt:
                due.append(c)
        if not due:
            return

        def probe(items=due):
            return [(c, yemot.find_scheduled(c.get("campaign_id"))) for c in items]

        self._sched_checker = _TaskWorker(probe, self)
        self._sched_checker.done.connect(self._on_sched_checked)
        self._sched_checker.start()

    def _on_sched_checked(self, res):
        self._sched_checker = None
        if isinstance(res, Exception):
            return                       # transient — retried on the next refresh
        track = None                     # (guid, campaignId, total)
        changed = False
        zone = timefmt._israel_zone()
        now = datetime.now(zone) if zone is not None else datetime.now().astimezone()
        for camp, (state, rec) in res:
            if state == "pending":
                continue                 # the server is a little behind — wait
            if state == "successful":
                cid = ""
                for key in ("campaignId", "campaign_id", "runId"):
                    cid = str(yemot._dig(rec, key) or "").strip()
                    if cid:
                        break
                # sent_at stays the PLANNED time — that is when the server
                # dialed, not the moment this app happened to notice.
                if cid:
                    db.update_tzintuk_campaign(camp["guid"], 0, 0, "sending",
                                               campaign_id=cid)
                    track = (camp["guid"], cid, int(camp.get("total") or 0))
                else:
                    # No campaign id to poll → close it as done (unknown
                    # results) rather than leaving a 'sending' record that
                    # polls a schedId forever.
                    db.update_tzintuk_campaign(camp["guid"], 0, 0, "done")
                changed = True
            elif state == "failed":
                db.update_tzintuk_campaign(camp["guid"], 0, 0, "sched_failed")
                changed = True
            else:                        # missing on the server
                dt = self._sched_dt(camp)
                if dt is not None and now - dt > timedelta(hours=1):
                    db.update_tzintuk_campaign(camp["guid"], 0, 0, "sched_failed")
                    changed = True
        if changed:
            self._refresh_history()
            self._refresh_sched_banner()
        if track and self._worker is None:
            self._active_guid = track[0]
            self._start_tracking(track[1], track[2])

    # ── Live tracking ─────────────────────────────────────────────────────────

    def _start_tracking(self, campaign_id: str, total: int, since_iso: str = ""):
        self.prog_frame.setVisible(True)
        self.btn_resend.setVisible(False)
        self.lbl_prog.setText("שולח בזמן אמת… אפשר להמשיך לעבוד, אל תסגור את התוכנה")
        self.progress.setRange(0, max(1, total))
        self.progress.setValue(0)
        self.lbl_ans.setText(self._answers_html(None))
        self.lbl_done.setText("הצליחו 0")
        self.lbl_fail.setText("נכשלו 0")
        self.lbl_wait.setText(f"ממתינים {total}")
        self._worker = _PollWorker(campaign_id, self, since_iso)
        self._worker.tick.connect(self._on_tick)
        self._worker.finished.connect(self._on_worker_done)
        self._worker.start()
        self._update_metrics()

    def _on_tick(self, st):
        if isinstance(st, Exception):
            self.lbl_prog.setText("החיבור למעקב נכשל — הקמפיין ממשיך לרוץ בימות; "
                                  "התוצאות יתעדכנו בכניסה הבאה ללשונית.")
            return
        done = st["delivered"] + st["failed"]
        self.progress.setRange(0, max(1, st["total"]))
        self.progress.setValue(min(done, st["total"]))
        self.lbl_ans.setText(self._answers_html(st.get("answers"), final=bool(st.get("finished"))))
        self.lbl_done.setText(f"הצליחו {st['delivered']}")
        self.lbl_fail.setText(f"נכשלו {st['failed']}")
        self.lbl_wait.setText(f"ממתינים {st['pending']}")
        # The DB (and the sync journal) get ONE update — at the end. Journaling
        # every 4-second tick would spam the shared Drive folder for nothing;
        # the live numbers live in this strip until then.
        if st.get("finished") and self._active_guid:
            db.update_tzintuk_campaign(
                self._active_guid, st["delivered"], st["failed"], "done",
                json.dumps(st.get("entries") or [], ensure_ascii=False))
        if st.get("finished"):
            self._last_failed = [e for e in st.get("entries") or [] if e.get("failed")]
            self.lbl_prog.setText(
                f"הקמפיין הסתיים ✓ — {st['delivered']} קיבלו את ההודעה, "
                f"{st['failed']} נכשלו. התשובות בסקר (הקשה {yemot.SURVEY_EXT}) "
                "ממשיכות להתעדכן — כפתור \"רענן תשובות\" בהיסטוריה.")
            self.btn_resend.setText(f"🔄 שלח שוב ל-{len(self._last_failed)} שנכשלו")
            self.btn_resend.setVisible(bool(self._last_failed))
            self._last_entries = list(st.get("entries") or [])
            self._apply_results_to_table(self._last_entries, final=True)
            self._refresh_history()
            # The confirmation badges on the "חלוקה ורישום" list come from the
            # stored report — repaint it so they appear without a tab switch.
            gt = getattr(self.main, "group_tab", None)
            if gt is not None:
                try:
                    gt._populate()
                except Exception:
                    pass

    _ANSWER_STYLE = {"1": ("✓", "#166534"), "2": ("✗", "#b91c1c"), "3": ("?", "#b45309")}

    def _apply_results_to_table(self, entries, final: bool = False):
        """Write each person's final result into the status column, so the
        operator sees what each family ANSWERED on the survey (1/2/3), who
        just got the call, and who failed. A row may have several dialed
        numbers (#gaira) — the BEST outcome among them wins. `final` = the
        campaign is over, so "no survey answer" is shown as "לא הגיב"."""
        by_phone = {e.get("phone"): e for e in entries if e.get("phone")}
        labels = yemot.answer_labels()
        self.table.blockSignals(True)
        for i, row in enumerate(self._rows):
            mine = [by_phone[p] for p in row.get("send") or [] if p in by_phone]
            if not mine or i >= self.table.rowCount():
                continue
            statuses = {str(e.get("status") or "") for e in mine}
            answer = next((str(e.get("answer")) for e in mine
                           if str(e.get("answer") or "") in self._ANSWER_STYLE), "")
            checked = any("answer" in e for e in mine)
            if answer:
                mark, col = self._ANSWER_STYLE[answer]
                txt, color = f"{mark} {labels.get(answer, '')}", QColor(col)
            elif any(e.get("confirmed") for e in mine):    # legacy key-7 reports
                txt, color = f"✓ {labels['1']}", QColor("#166534")
            elif final and checked:
                txt, color = f"לא הגיב (לא הקיש {yemot.SURVEY_EXT})", QColor("#6b7280")
            elif "callback" in statuses:     # v2.96 — classic: called back
                txt, color = "📞 חזר לשיחה ושמע", QColor("#0f6e56")
            elif any(e.get("ok") for e in mine):
                txt, color = "קיבל את ההודעה", QColor("#0f6e56")
            elif all(e.get("failed") for e in mine):
                txt, color = "⚠ לא נענה / נכשל", QColor("#a32d2d")
            elif statuses and statuses <= {"no_callback"}:
                txt, color = "לא חזר לשיחה", QColor("#6b7280")
            else:
                continue
            it = QTableWidgetItem(txt)
            it.setForeground(color)
            self.table.setItem(i, 3, it)
        self.table.blockSignals(False)

    def _on_worker_done(self):
        self._worker = None
        self._update_metrics()

    # ── Classic-tzintuk callback watch (v2.96) ────────────────────────────────

    def _start_callback_tracking(self, guid: str, targets: dict,
                                 deadline: float, seed_entries=None,
                                 since_iso: str = ""):
        """Open the live watch window after a classic tzintuk: who calls the
        line back is caught in real time (Yemot keeps no call log, so this
        works only while the window is open); what they answered on the
        survey extension is read from its data file (persistent)."""
        if self._cb_worker is not None:
            return
        self._active_guid = guid
        self.prog_frame.setVisible(True)
        self.btn_resend.setVisible(False)
        self.btn_extend_track.setVisible(True)
        self.btn_stop_track.setVisible(True)
        self.progress.setRange(0, 0)          # time window — busy stripe
        self.lbl_fail.setText("")
        self.lbl_ans.setText(self._answers_html(None))
        self.lbl_done.setText("📞 חזרו לשיחה 0")
        self.lbl_wait.setText(f"טרם חזרו {len(targets)}")
        self._cb_last_persist = time.time()
        self._cb_worker = _CallbackWorker(targets, deadline, seed_entries, self, since_iso)
        self._cb_worker.tick.connect(self._on_cb_tick)
        self._cb_worker.finished.connect(self._on_cb_worker_done)
        self._cb_worker.start()
        self._update_metrics()

    def _extend_tracking(self):
        if self._cb_worker is not None:
            self._cb_worker.extend(30 * 60)

    def _stop_tracking_now(self):
        if self._cb_worker is not None:
            self._cb_worker.stop()     # the worker emits a final snapshot

    def _persist_callback(self, st, final: bool):
        if not self._active_guid:
            return
        db.update_tzintuk_campaign(
            self._active_guid, int(st.get("returned") or 0), 0,
            "done" if final else "sending",
            json.dumps(st.get("entries") or [], ensure_ascii=False))
        self._cb_last_persist = time.time()

    def _on_cb_tick(self, st):
        if not isinstance(st, dict):
            return
        entries = st.get("entries") or []
        returned = int(st.get("returned") or 0)
        self.lbl_ans.setText(self._answers_html(st.get("answers"), final=bool(st.get("done"))))
        self.lbl_done.setText(f"📞 חזרו לשיחה {returned}")
        self.lbl_wait.setText(f"טרם חזרו {max(0, len(entries) - returned)}")
        if st.get("done"):
            self.progress.setRange(0, 1)
            self.progress.setValue(1)
            self.lbl_prog.setText(
                f"המעקב הסתיים ✓ — {returned} חזרו לשיחה ושמעו את ההודעה. "
                f"התשובות בסקר (הקשה {yemot.SURVEY_EXT}) נשמרות בקו וממשיכות "
                "להתעדכן — כפתור \"רענן תשובות\" בהיסטוריה.")
            self._persist_callback(st, final=True)
            self._last_entries = list(entries)
            self._apply_results_to_table(entries, final=True)
            self._refresh_history()
            gt = getattr(self.main, "group_tab", None)
            if gt is not None:
                try:
                    gt._populate()   # תגי "אישר הגעה" ברשימת החלוקה
                except Exception:
                    pass
            return
        mins = int(st.get("remaining") or 0) // 60
        base = ("📞 צינתוק קלאסי — עוקב אחרי מי שמתקשר חזרה לקו "
                f"(עוד ~{max(1, mins)} דק'). אפשר להמשיך לעבוד, "
                "אל תסגור את התוכנה.")
        if st.get("error"):
            base += "  ⚠ תקלת תקשורת זמנית במעקב — ממשיך לנסות."
        self.lbl_prog.setText(base)
        if st.get("changed"):
            # in-window: repaint only who RETURNED (the rest stay "מוכן")
            self._last_entries = [e for e in entries if e.get("ok")]
            self._apply_results_to_table(self._last_entries)
            if time.time() - self._cb_last_persist > 120:
                self._persist_callback(st, final=False)

    def _on_cb_worker_done(self):
        self._cb_worker = None
        self.btn_extend_track.setVisible(False)
        self.btn_stop_track.setVisible(False)
        self._update_metrics()

    def _resume_classic(self, camp: dict):
        """A classic tzintuk is still 'sending' (the app closed mid-window, or
        the other computer sent it): reopen the watch if the window is still
        open; otherwise freeze what was collected as the final result."""
        try:
            entries = json.loads(camp.get("report_json") or "[]")
        except ValueError:
            entries = []
        entries = [e for e in entries if isinstance(e, dict)]
        targets = {e.get("phone"): e.get("name") or ""
                   for e in entries if e.get("phone")}
        sent = timefmt.to_israel(camp.get("sent_at") or "")
        deadline = (sent.timestamp() + yemot.CLASSIC_TRACK_SECONDS
                    if sent is not None else 0.0)
        if targets and deadline > time.time() + 5:
            self._start_callback_tracking(camp["guid"], targets, deadline,
                                          seed_entries=entries,
                                          since_iso=camp.get("sent_at") or "")
        elif deadline and time.time() - deadline > 600:
            # long past (10-min grace for the sender's own finalize) — close it
            returned = sum(1 for e in entries if e.get("ok"))
            db.update_tzintuk_campaign(
                camp["guid"], returned, 0, "done",
                json.dumps(entries, ensure_ascii=False))
            self._refresh_history()

    def _maybe_resume_tracking(self):
        """The app (or the tab) was closed mid-campaign: the newest record is
        still 'sending'. Pick its tracking back up so the results get written —
        works also for a campaign the OTHER computer started (same account)."""
        self._check_scheduled()          # #xi85i — due schedules first
        self._sync_history(manual=False)  # v3.10 — server history, once a day
        if (self._worker is not None or self._cb_worker is not None
                or not yemot.is_configured()):
            return
        camps = [c for c in db.get_tzintuk_campaigns(limit=5)
                 if c.get("status") != "scheduled"][:1]
        if not camps or camps[0].get("status") != "sending":
            self._auto_refresh_answers()   # v3.02 — late survey answers
            return
        camp = camps[0]
        if (camp.get("name") or "").startswith("צינתוק קלאסי"):
            self._resume_classic(camp)   # v2.96 — callback watch, not polling
            return
        if camp.get("campaign_id"):
            self._active_guid = camp["guid"]
            self._start_tracking(camp["campaign_id"],
                                 int(camp.get("total") or 0),
                                 camp.get("sent_at") or "")

    # ── Survey answers (v3.02) ────────────────────────────────────────────────

    @staticmethod
    def _answers_html(counts, final: bool = False) -> str:
        """Rich-text counters for the tracking strip:
        '✓ מגיע 2 · ✗ לא מגיע 1 · ? לא יודע 0 · לא הגיבו 5' (the last part
        only once the campaign is over)."""
        labels = yemot.answer_labels()
        c = counts or {k: 0 for k in yemot.ANSWER_KEYS}
        parts = [f"<span style='color:{col}'>{mark} {labels[k]} {int(c.get(k, 0))}</span>"
                 for k, (mark, col) in TzintukimTab._ANSWER_STYLE.items()]
        if final and counts is not None:
            parts.append(f"<span style='color:#6b7280'>לא הגיבו {int(c.get('', 0))}</span>")
        return " · ".join(parts)

    def _answer_campaigns(self) -> list:
        """Finished campaigns from the last 14 days — the ones whose survey
        answers can still change."""
        out = []
        for c in db.get_tzintuk_campaigns(limit=30):
            if c.get("status") != "done":
                continue
            sent = timefmt.to_israel(c.get("sent_at") or "")
            if sent is None or time.time() - sent.timestamp() > 14 * 86400:
                continue
            out.append(c)
        return out

    # ── Server history (v3.10) ───────────────────────────────────────────────

    def _refresh_hist_sync_label(self):
        if self._hist_worker is not None:
            self.lbl_hist_sync.setText("⏳ " + (self._hist_prog or "מושך היסטוריה מהשרת…"))
            return
        s = call_history.summary()
        if not s["campaigns"] and not s["calls"]:
            self.lbl_hist_sync.setText(
                "ידע על שעות מענה: עדיין לא נמשכה היסטוריה מהשרת של ימות.")
            return
        when = timefmt.relative(s["updated"]) if s["updated"] else ""
        self.lbl_hist_sync.setText(
            f"ידע על שעות מענה מהשרת: {s['campaigns']} קמפיינים ו-{s['calls']:,} "
            f"שיחות נכנסות לקו ({len(s['months'])} חודשים)"
            + (f" · עודכן {when}" if when else ""))

    def _sync_history(self, manual: bool):
        """Pull the Yemot server's campaign + incoming-call history into the
        local cache (utils.call_history) in the background. Automatic once a
        day; the button forces it and reports at the end."""
        if self._hist_worker is not None or not yemot.is_configured():
            if manual and self._hist_worker is not None:
                QMessageBox.information(self, "צינתוקים", "המשיכה מהשרת כבר רצה.")
            return
        if not manual and not call_history.is_stale():
            return
        self._hist_prog = ""

        def _prog(text):
            self._hist_prog = text

        self._hist_worker = _TaskWorker(
            lambda: call_history.sync_from_server(progress=_prog), self)
        timer = QTimer(self)
        timer.setInterval(700)
        timer.timeout.connect(self._refresh_hist_sync_label)

        def _done(res):
            timer.stop()
            self._hist_worker = None
            try:
                self._stats = yemot.answer_stats()
                if self._list_loaded:
                    self._populate()
            except Exception:
                pass
            self._refresh_hist_sync_label()
            if not manual:
                return
            if isinstance(res, Exception):
                QMessageBox.warning(self, "צינתוקים",
                                    f"משיכת ההיסטוריה מהשרת נכשלה:\n{res}")
                return
            errs = int(res.get("errors") or 0)
            QMessageBox.information(
                self, "צינתוקים",
                f"ההיסטוריה עודכנה ✓\n"
                f"קמפיינים חדשים: {res.get('new_campaigns', 0)} · "
                f"חודשי יומן שנקראו: {res.get('months_fetched', 0)}\n"
                f"סה\"כ בידע: {res.get('campaigns', 0)} קמפיינים, "
                f"{res.get('calls', 0):,} שיחות נכנסות"
                + (f"\n({errs} פריטים לא נקראו — ינוסו שוב בפעם הבאה)" if errs else ""))
        self._hist_worker.done.connect(_done)
        self.btn_hist_sync.setEnabled(False)
        self._hist_worker.finished.connect(lambda: self.btn_hist_sync.setEnabled(True))
        timer.start()
        self._hist_worker.start()
        self._refresh_hist_sync_label()

    def _refresh_answers(self):
        """Button: read the survey extension's answers now and refresh the
        table, the history and the distribution list."""
        if not self._require_config():
            return
        if not self._answer_campaigns():
            QMessageBox.information(self, "צינתוקים",
                                    "אין צינתוק מהשבועיים האחרונים לעדכן.")
            return
        try:
            rows = self._run_blocking(yemot.fetch_survey_rows, "קורא את התשובות מהקו…")
        except Exception as e:
            QMessageBox.warning(self, "צינתוקים", f"קריאת התשובות נכשלה:\n{e}")
            return
        changed = self._apply_answer_rows(rows)
        QMessageBox.information(
            self, "צינתוקים",
            "התשובות עודכנו ✓" if changed else "אין תשובות חדשות מאז הפעם הקודמת.")

    def _auto_refresh_answers(self):
        """Background refresh (from refresh()/resume) — at most once per 10
        minutes, only when nothing else is polling the server."""
        if (self._worker is not None or self._cb_worker is not None
                or getattr(self, "_ans_worker", None) is not None
                or not yemot.is_configured()
                or time.time() - getattr(self, "_ans_last", 0.0) < 600
                or not self._answer_campaigns()):
            return
        self._ans_last = time.time()
        self._ans_worker = _TaskWorker(yemot.fetch_survey_rows, self)

        def _done(res):
            self._ans_worker = None
            if isinstance(res, list):
                try:
                    self._apply_answer_rows(res)
                except Exception:
                    pass
        self._ans_worker.done.connect(_done)
        self._ans_worker.start()

    def _apply_answer_rows(self, rows) -> bool:
        """Merge survey rows into every recent finished campaign; persist only
        the ones that changed (each write is a sync-journal entry). Repaints
        the loaded list when it belongs to the newest of them."""
        changed_any = False
        newest = None
        for camp in self._answer_campaigns():
            entries = yemot._report_entries(camp)
            if not entries:
                continue
            entries, changed = yemot.merge_survey_answers(entries, rows, camp.get("sent_at") or "")
            if changed:
                db.update_tzintuk_campaign(
                    camp["guid"], int(camp.get("delivered") or 0),
                    int(camp.get("failed") or 0), "done",
                    json.dumps(entries, ensure_ascii=False))
                changed_any = True
            if newest is None:
                newest = (camp, entries)
        if newest is not None and self._list_loaded:
            camp, entries = newest
            if (camp.get("dist_date") or "") == self._dist_date_iso():
                self._last_entries = list(entries)
                self._apply_results_to_table(entries, final=True)
        self._refresh_history()
        gt = getattr(self.main, "group_tab", None)
        if gt is not None and changed_any:
            try:
                gt._populate()
            except Exception:
                pass
        return changed_any

    # ── History ───────────────────────────────────────────────────────────────

    def _refresh_history(self):
        camps = db.get_tzintuk_campaigns(limit=100)
        self.hist.setRowCount(len(camps))
        status_he = {"sending": "בתהליך", "done": "הסתיים",
                     "scheduled": "מתוזמן ⏳", "canceled": "בוטל",
                     "sched_failed": "התזמון נכשל"}
        for i, c in enumerate(camps):
            when = timefmt.datetime_str(c.get("sent_at") or "")
            src = c.get("device") or ""
            self.hist.setItem(i, 0, QTableWidgetItem(when + (f"  ({src})" if src else "")))
            name = c.get("name") or ""
            st = status_he.get(c.get("status") or "", c.get("status") or "")
            self.hist.setItem(i, 1, QTableWidgetItem(f"{name} — {st}"))
            self.hist.setItem(i, 2, QTableWidgetItem(str(c.get("total") or 0)))
            self.hist.setItem(i, 3, QTableWidgetItem(str(c.get("delivered") or 0)))
            ans = QTableWidgetItem(self._answers_text(c))
            ans.setForeground(QColor("#166534"))
            self.hist.setItem(i, 4, ans)
            self.hist.setItem(i, 5, QTableWidgetItem(str(c.get("failed") or 0)))

    def _export_history(self):
        """#67rdi — ייצוא כל היסטוריית הצינתוקים לאקסל, כולל סטטוס פר-מספר
        (מה ענה בסקר — מגיע / לא מגיע / לא יודע / לא הגיב — מי רק קיבל ומי לא נענה)."""
        from utils.ui import reveal_in_folder
        from utils.excel_utils import export_tzintuk_history_to_excel
        camps = db.get_tzintuk_campaigns()
        if not camps:
            QMessageBox.information(self, "צינתוקים", "עדיין אין צינתוקים בהיסטוריה.")
            return
        # שם לכל מספר — לדוחות ישנים שבהם ימות לא החזיר את השם.
        name_by_phone = {}
        try:
            for rec in db.get_all_recipients():
                nm = rec.get("full_name") or ""
                for p in yemot.pick_phones(rec):
                    name_by_phone.setdefault(p, nm)
        except Exception:
            pass
        try:
            with busy_cursor():
                path = export_tzintuk_history_to_excel(camps, name_by_phone)
            reveal_in_folder(path)
            QMessageBox.information(self, "ייצוא הושלם",
                                    f"הקובץ נשמר ונפתחה התיקייה:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "שגיאה", str(e))

    @staticmethod
    def _answers_text(camp: dict) -> str:
        """'✓2 ✗1 ?0 · 5 לא הגיבו' — derived from the stored report (no DB
        column; the report is synced so both computers agree). Old reports
        (before the survey) show the legacy key-7 count."""
        entries = yemot._report_entries(camp)
        if not yemot.survey_checked(entries):
            legacy = sum(1 for e in entries if e.get("confirmed")
                         or str(e.get("status") or "").lower() == "accepted")
            return f"✓{legacy}" if legacy else "—"
        a = yemot.answer_counts(entries)
        return f"✓{a['1']} ✗{a['2']} ?{a['3']} · {a['']} לא הגיבו"
