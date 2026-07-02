# -*- coding: utf-8 -*-
"""Round-trip test for the volunteer-checklist feature: export a checklist,
simulate a volunteer marking who came + notes, import it back, and verify the
distributions land in history correctly. Touches only a TEMP db."""
import os, sys, tempfile
os.environ["PYTHONUTF8"] = "1"
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
import database as db
db.DB_PATH = os.path.join(tempfile.gettempdir(), "vol_flow_test.db")
db.BACKUP_DIR = os.path.join(tempfile.gettempdir(), "vol_flow_test_bk")
for e in ("", "-wal", "-shm"):
    try: os.remove(db.DB_PATH + e)
    except OSError: pass
db.init_db()

import openpyxl
from utils.excel_utils import (export_volunteer_checklist_to_excel, import_volunteer_checklist,
                               _VOL_DATA_START_ROW, _VOL_COL_CAME, _VOL_COL_NOTE,
                               _VOL_COL_ID, _VOL_GENERAL_NOTE_ROW)
from utils import email_utils

ok = True
def check(label, cond, extra=""):
    global ok; ok = ok and cond
    print(f"  [{'OK' if cond else 'FAIL'}] {label}" + (f"  {extra}" if extra else ""))

print("=== seed recipients ===")
ids = []
for i in range(5):
    rid = db.add_recipient({"full_name": f"נבדק {i}", "status": "פעיל", "frequency": "שבועי",
                            "area": "בעלז", "souls": 3 + i, "phone1": f"05{i}1111111"})
    ids.append(rid)
recs = [db.get_recipient(i) for i in ids]

print("\n=== export volunteer checklist ===")
path = export_volunteer_checklist_to_excel(recs, "2026-07-08", "בדיקת מתנדב",
                                           "סל מזון", 25, "המתנדב הבודק")
check("file created", os.path.exists(path))
check("filename is xlsx", path.endswith(".xlsx"))

wb = openpyxl.load_workbook(path)
check("hidden meta sheet present", "meta" in wb.sheetnames)
check("meta sheet actually hidden", wb["meta"].sheet_state == "hidden")
ws = wb.worksheets[0]
check("id column hidden", ws.column_dimensions[
    openpyxl.utils.get_column_letter(_VOL_COL_ID)].hidden is True)
check("display title uses dd/mm/yyyy", "08/07/2026" in str(ws["A2"].value))

print("\n=== simulate volunteer filling it in ===")
# 3 came (rows 0-2), 1 explicitly did not (row 3), 1 left blank (row 4)
for i in range(3):
    ws.cell(_VOL_DATA_START_ROW + i, _VOL_COL_CAME, "כן")
ws.cell(_VOL_DATA_START_ROW, _VOL_COL_NOTE, "קיבל תוספת חלב")
ws.cell(_VOL_DATA_START_ROW + 3, _VOL_COL_CAME, "לא")
ws.cell(_VOL_GENERAL_NOTE_ROW, 2, "החלוקה עברה בשלום, חסר סל אחד בסוף")
wb.save(path)

print("\n=== import it back ===")
result = import_volunteer_checklist(path)
check("3 received rows parsed", len(result["received"]) == 3, f"got {len(result['received'])}")
check("no unmatched rows", len(result["unmatched"]) == 0, str(result["unmatched"]))
check("meta dist_date is ISO", result["meta"]["dist_date"] == "2026-07-08")
check("meta what/distributor/dist_name correct",
      result["meta"]["what"] == "סל מזון" and result["meta"]["distributor"] == "המתנדב הבודק"
      and result["meta"]["dist_name"] == "בדיקת מתנדב")
check("meta qty is int", result["meta"]["qty"] == 25)
check("general note captured", "עברה בשלום" in result["meta"]["general_note"])
check("per-recipient note captured", any("תוספת חלב" in (r["notes"] or "") for r in result["received"]))
check("ids match originals", {r["id"] for r in result["received"]} == set(ids[:3]))

print("\n=== id-tampered fallback matches by name ===")
path2 = export_volunteer_checklist_to_excel(recs, "2026-07-08", "בדיקת נפילה", "עוף", 10, "רותי")
wb2 = openpyxl.load_workbook(path2)
ws2 = wb2.worksheets[0]
ws2.cell(_VOL_DATA_START_ROW, _VOL_COL_ID, None)   # blank the hidden id
ws2.cell(_VOL_DATA_START_ROW, _VOL_COL_CAME, "כן")
wb2.save(path2)
result2 = import_volunteer_checklist(path2)
check("fallback-by-name recovered the row", len(result2["received"]) == 1
      and result2["received"][0]["id"] == ids[0])

print("\n=== bulk_add_distributions writes correct history ===")
n_before = len(db.get_distributions())
records = []
for r in result["received"]:
    rec = dict(r)
    rec["notes"] = (r.get("notes") or "") + f" | הערה כללית: {result['meta']['general_note']}"
    records.append(rec)
db.bulk_add_distributions(records, result["meta"]["dist_date"], result["meta"]["what"],
                          result["meta"]["qty"], result["meta"]["distributor"])
n_after = len(db.get_distributions())
check("3 distributions recorded", n_after - n_before == 3, f"delta={n_after - n_before}")
dist_rows = db.get_distributions(limit=3)
check("distributor name recorded correctly",
      all(d.get("distributor") == "המתנדב הבודק" for d in dist_rows))
check("general note suffix present in history",
      any("הערה כללית" in (d.get("notes") or "") for d in dist_rows))
rec0 = db.get_recipient(ids[0])
check("recipient last_distribution updated", rec0.get("last_distribution") == "2026-07-08")

print("\n=== email module builds messages without a real send ===")
check("not configured returns clear error",
      not email_utils.is_configured())
try:
    email_utils.send_email("x@example.com", "subj", "<p>hi</p>", attachment_path=path)
    check("send_email without config raises", False)
except RuntimeError as e:
    check("send_email without config raises RuntimeError", "לא הוגדרו" in str(e))

for p in (path, path2):
    try: os.remove(p)
    except OSError: pass

print("\n" + "=" * 55)
print("RESULT:", "ALL PASS ✓" if ok else "FAILURES ✗")
sys.exit(0 if ok else 1)
