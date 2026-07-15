import os
import re
import sys
from datetime import datetime, date as _date_type
from pathlib import Path
from typing import List, Dict


def _app_dir() -> Path:
    """Return the app root directory — works in dev and as PyInstaller EXE."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parents[1]


def _resource_path(rel: str) -> str:
    """Locate a bundled resource (e.g. org_logo.png) in both dev and frozen mode."""
    base = getattr(sys, "_MEIPASS", str(_app_dir()))
    return os.path.join(base, rel)


def _checklist_password() -> str:
    """The open-password for the volunteer checklist file (empty = none)."""
    try:
        from utils import email_utils
        return email_utils.get_checklist_password()
    except Exception:
        return ""


def _encrypt_xlsx(path: str, password: str):
    """Encrypt an .xlsx in place so it requires `password` to open in Excel.
    No-op when password is empty. Uses msoffcrypto (pure-Python OOXML agile
    encryption) — the same standard format Excel's own 'Encrypt with Password'
    produces, so any Excel/LibreOffice/phone viewer prompts for it normally."""
    if not password:
        return
    import io
    import msoffcrypto
    with open(path, "rb") as f:
        plain = io.BytesIO(f.read())
    off = msoffcrypto.OfficeFile(plain)
    with open(path, "wb") as out:
        off.encrypt(password, out)


def _load_maybe_encrypted(path: str, **kw):
    """openpyxl.load_workbook that transparently opens password-encrypted files.
    Tries the file as-is first; if that fails and a checklist password is set,
    decrypts in memory and loads that. Raises if it can't be opened either way."""
    import openpyxl
    try:
        return openpyxl.load_workbook(path, **kw)
    except Exception:
        pw = _checklist_password()
        if not pw:
            raise
        import io
        import msoffcrypto
        dec = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=pw)
            off.decrypt(dec)
        dec.seek(0)
        return openpyxl.load_workbook(dec, **kw)


def _downloads_dir() -> Path:
    """The user's Downloads folder (where all exports go). Falls back to an
    'exports' folder next to the app if Downloads can't be resolved/created."""
    try:
        d = Path(os.path.expanduser("~")) / "Downloads"
        d.mkdir(parents=True, exist_ok=True)
        return d
    except Exception:
        d = _app_dir() / "exports"
        d.mkdir(parents=True, exist_ok=True)
        return d


def _parse_date(val) -> str:
    """Convert any date representation to ISO yyyy-mm-dd string."""
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, _date_type):
        return val.isoformat()
    val = str(val).strip()
    if not val or val == "None":
        return ""
    # yyyy-mm-dd or yyyy-mm-dd HH:MM:SS
    if len(val) >= 10 and val[4] == "-" and val[7] == "-":
        return val[:10]
    for fmt in ("%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(val, fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def _normalize_phone(val: str) -> str:
    """Add leading 0 to 9-digit Israeli mobile numbers that lost it in Excel."""
    if val and val.isdigit() and len(val) == 9 and val[0] == "5":
        return "0" + val
    return val


def _write_phone_cell(cell, value):
    """Write a phone number so Excel doesn't flag it with the green-triangle
    'מספר המאוחסן כטקסט' (number-stored-as-text) warning. A pure-digit phone is
    stored as a real number with a leading-zero-preserving format (so 0533193925
    stays exactly that, no warning). Anything with non-digits — a landline hyphen
    (04-9955317), '+', spaces — stays plain text; Excel doesn't flag those, since
    they aren't numeric."""
    s = str(value if value is not None else "").strip()
    if s.isdigit() and 1 <= len(s) <= 15:
        cell.value = int(s)
        cell.number_format = "0" * len(s)   # keep exact length incl. leading zero
    else:
        cell.value = s


# Hebrew final-form letters → their regular forms, so a substring match isn't
# defeated by ן↔נ etc. (e.g. "אלמן" ends in a final nun but "אלמנה" uses a
# regular nun — without this they wouldn't match).
_HEB_FINALS = str.maketrans("ךםןףץ", "כמנפצ")

# Marital-status roots (final letters normalized) that mean a ONE-adult
# household. Matched as substrings so masculine/feminine spellings are both
# covered (גרוש/גרושה, אלמן/אלמנה, רווק/רווקה, פרוד/פרודה).
_SINGLE_PARENT_MARKERS = ("גרוש", "אלמנ", "רווק", "פרוד", "חד הורי", "חד-הורי", "לא נשוי")


def _adults_in_household(marital_status) -> int:
    """How many adults to count toward נפשות for this household. A single-parent
    status (divorced/widowed/single/separated) → 1 adult; anything else,
    including a blank/unknown status (the common case), → 2 (a couple)."""
    s = str(marital_status or "").strip().translate(_HEB_FINALS)
    return 1 if any(m in s for m in _SINGLE_PARENT_MARKERS) else 2


def import_from_excel(path: str) -> List[Dict]:
    """Read recipients from תבנית ליהודה format Excel file."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row: look for template-specific columns
    header_row_idx = None
    header = []
    for i, row in enumerate(rows):
        texts = [str(c).strip() if c else "" for c in row]
        if any(t in ("משפחה", "מספר מזהה", "נייד בעל") for t in texts):
            header = texts
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError(
            "לא נמצאה שורת כותרת בפורמט תבנית ליהודה.\n"
            "הקובץ חייב להכיל עמודות: 'משפחה', 'פרטי', 'נייד בעל' וכו'."
        )

    # Map column indices — explicit template column names only
    col_map = {}
    for idx, h in enumerate(header):
        h_s = h.strip() if h else ""
        if not h_s:
            continue
        if h_s == "מספר מזהה":
            col_map["external_id"] = idx
        elif h_s == "מקור":
            col_map["source"] = idx
        elif h_s == "משפחה":
            col_map["family_name"] = idx
        elif h_s == "פרטי":
            col_map.setdefault("first_name", idx)
        elif h_s in ("ת. לידה", "ת.לידה") and "birth_date" not in col_map:
            col_map["birth_date"] = idx
        elif h_s in ("ת. לידה בן /בת הזוג", "ת. לידה בן/בת הזוג"):
            col_map["spouse_birth_date"] = idx
        elif h_s == "תז בעל":
            col_map["id_number"] = idx
        elif h_s == "תז אשה":
            col_map["spouse_id_number"] = idx
        # "ללא מקף" columns have clean digits — prefer them for phone storage
        elif h_s == "נייד בעל ללא מקף":
            col_map.setdefault("phone1", idx)
        elif h_s == "נייד אשה ללא מקף":
            col_map.setdefault("phone2", idx)
        elif h_s == "נייד בעל":
            col_map.setdefault("phone1", idx)
        elif h_s == "נייד אשה":
            col_map.setdefault("phone2", idx)
        elif h_s == "טלפון בבית":
            col_map.setdefault("phone3", idx)
        elif h_s == "טלפון נוסף":
            col_map.setdefault("phone3", idx)
        elif h_s == "ילדים בבית":
            col_map["children_home"] = idx
        elif h_s == "ילדים נשואים":
            col_map["children_married"] = idx
        elif h_s == "מספר ילדים":
            col_map["children_total"] = idx
        elif h_s == "מצב אישי":
            col_map["marital_status"] = idx
        elif h_s in ("אימייל", "email", "Email"):
            col_map["email"] = idx
        elif h_s == "בית כנסת":
            col_map["synagogue"] = idx
        elif h_s == "הוצאות דיור":
            col_map["housing_expenses"] = idx
        elif h_s == "הוצאות רפואיות":
            col_map["medical_expenses"] = idx
        elif h_s == "הכנסות":
            col_map["income"] = idx
        elif h_s == "פנוי לנפש":
            col_map["per_soul"] = idx
        elif "הקף משרה" in h_s or "היקף משרה" in h_s:
            col_map["work_scope"] = idx
        elif h_s == "סוג הורה":
            col_map["parent_type"] = idx
        elif h_s == "עיסוק בעל":
            col_map["occupation"] = idx
        elif h_s == "קהילה":
            col_map["area"] = idx
        elif h_s == "רחוב":
            col_map["street"] = idx
        elif h_s == "מספר" and "house_num" not in col_map:
            col_map["house_num"] = idx
        elif h_s == "שם נציג":
            col_map["representative"] = idx

    # Priority code column — unnamed, sits right before "משפחה". Holds the
    # distribution code (e.g. 4=קבוע, 3=עדיפות ראשונה, 2=שנייה, חובת בירור...).
    priority_idx = None
    _fam_idx = col_map.get("family_name")
    if _fam_idx is not None and _fam_idx - 1 >= 0 and not str(header[_fam_idx - 1] or "").strip():
        priority_idx = _fam_idx - 1

    results = []
    for row in rows[header_row_idx + 1:]:
        if not any(row):
            continue

        def cell(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        def raw_cell(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        def _int_cell(key):
            v = cell(key)
            try:
                return int(float(v)) if v else 0
            except (ValueError, TypeError):
                return 0

        # Name: combine משפחה + פרטי
        family = cell("family_name")
        first  = cell("first_name")
        name   = (family + " " + first).strip() if (family or first) else ""
        if not name or name in ("None", "0"):
            continue

        # Address: combine רחוב + מספר
        street = cell("street")
        house  = cell("house_num")
        address = (street + " " + house).strip() if street else house

        # Souls: children at home + the adults in the household. A single-parent
        # household (divorced/widowed/single) has ONE adult, not two — counting a
        # second adult would over-state נפשות and skew the need-score. A blank or
        # unknown status defaults to two (a couple — the common case).
        children_home = _int_cell("children_home")
        marital = cell("marital_status")
        souls = children_home + _adults_in_household(marital)

        # Priority code → priority number + frequency.
        # 4 = קבוע → weekly regular flow; 3/2/1/0 = one-time (priority tiers);
        # 'חובת בירור' (no digit) = one-time, kept as data only. The V/X letters
        # are ignored — only the number matters.
        priority = None
        priority_raw = ""
        if priority_idx is not None and priority_idx < len(row):
            rawv = row[priority_idx]
            if rawv is not None:
                priority_raw = str(rawv).strip()
                m = re.search(r"\d+", priority_raw)
                if m:
                    priority = int(m.group())
        if priority == 4:
            frequency = "שבועי"
        elif priority in (2, 3) or "בירור" in priority_raw:
            frequency = "חד-פעמי"
        else:
            # Codes 1/0, stray non-priority marks (V/X/text), or nothing at all
            # carry no real distribution meaning — don't invent 'חד-פעמי' for
            # them. Take the file's own frequency column if present, otherwise
            # leave it blank (an unmarked recipient stays unmarked).
            frequency = cell("frequency")

        results.append({
            "full_name":         name,
            "phone1":            _normalize_phone(cell("phone1")),
            "phone2":            _normalize_phone(cell("phone2")),
            "phone3":            _normalize_phone(cell("phone3")),
            "address":           address,
            "area":              cell("area"),
            "souls":             souls,
            "frequency":         frequency,
            "priority":          priority,
            "priority_raw":      priority_raw,
            "status":            cell("status") or "פעיל",
            "last_distribution": _parse_date(raw_cell("last_distribution")),
            "next_distribution": _parse_date(raw_cell("next_distribution")),
            "start_date":        _parse_date(raw_cell("start_date")),
            "external_id":       cell("external_id"),
            "source":            cell("source"),
            "birth_date":        _parse_date(raw_cell("birth_date")),
            "spouse_birth_date": _parse_date(raw_cell("spouse_birth_date")),
            "id_number":         cell("id_number"),
            "spouse_id_number":  cell("spouse_id_number"),
            "children_home":     children_home,
            "children_married":  _int_cell("children_married"),
            "children_total":    _int_cell("children_total"),
            "marital_status":    marital,
            "email":             cell("email"),
            "synagogue":         cell("synagogue"),
            "housing_expenses":  cell("housing_expenses"),
            "medical_expenses":  cell("medical_expenses"),
            "income":            cell("income"),
            "per_soul":          cell("per_soul"),
            "work_scope":        cell("work_scope"),
            "parent_type":       cell("parent_type"),
            "occupation":        cell("occupation"),
            "representative":    cell("representative"),
        })

    return results


def export_distribution_to_excel(recipients: List[Dict], dist_date: str) -> str:
    """Export distribution list to Excel. Returns saved file path."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "רשימת חלוקה"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["מס'", "שם מלא", "טלפון 1", "טלפון 2", "טלפון 3", "אזור", "נפשות", "✓ ביצוע"]
    phone_cells = [(3, "phone1"), (4, "phone2"), (5, "phone3")]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[1].height = 22

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    normal_fill = PatternFill("solid", fgColor="FFFFFF")
    # Build the style objects ONCE and reuse them. Creating a fresh Alignment per
    # cell meant ~32k objects for a few-thousand-row export, which took seconds.
    cell_align = Alignment(horizontal="right", vertical="center")

    for i, rec in enumerate(recipients, 1):
        row = [i, rec.get("full_name", ""), rec.get("phone1", ""),
               rec.get("phone2", ""), rec.get("phone3", ""),
               rec.get("area", ""), rec.get("souls", ""), ""]
        ws.append(row)
        fill = alt_fill if i % 2 == 0 else normal_fill
        for cell in ws[i + 1]:
            cell.alignment = cell_align
            cell.fill = fill
            cell.border = border
        for c, key in phone_cells:   # real-number phones → no green-triangle warning
            _write_phone_cell(ws.cell(i + 1, c), rec.get(key, ""))
        ws.row_dimensions[i + 1].height = 18

    col_widths = [6, 26, 16, 16, 16, 10, 8, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    ws.insert_rows(1)
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"רשימת חלוקה — {dist_date}"
    title_cell.font = Font(bold=True, size=14, color="1D4ED8")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    exports_dir = _downloads_dir()
    filename = f"חלוקה_{dist_date.replace('/', '-')}_{datetime.now().strftime('%H%M%S')}.xlsx"
    path = str(exports_dir / filename)
    wb.save(path)
    return path


# Full recipient field set for the detailed export (key, Hebrew header).
_FULL_FIELDS = [
    ("full_name", "שם מלא"), ("priority", "עדיפות"),
    ("phone1", "טלפון 1"), ("phone2", "טלפון 2"), ("phone3", "טלפון 3"),
    ("address", "כתובת"), ("area", "אזור"), ("souls", "נפשות"),
    ("frequency", "תדירות"), ("last_distribution", "חלוקה אחרונה"),
    ("next_distribution", "חלוקה הבאה"), ("status", "סטטוס"), ("notes", "הערות"),
    ("external_id", "מס' מזהה"), ("source", "מקור"),
    ("birth_date", "ת. לידה"), ("spouse_birth_date", "ת. לידה בן/בת זוג"),
    ("id_number", "ת.ז. בעל"), ("spouse_id_number", "ת.ז. אשה"),
    ("children_home", "ילדים בבית"), ("children_married", "ילדים נשואים"),
    ("children_total", "מספר ילדים"), ("marital_status", "מצב אישי"),
    ("email", "אימייל"), ("synagogue", "בית כנסת"),
    ("housing_expenses", "הוצ' דיור"), ("medical_expenses", "הוצ' רפואיות"),
    ("income", "הכנסות"), ("per_soul", "פנוי לנפש"),
    ("work_scope", "היקף משרה"), ("parent_type", "סוג הורה"),
    ("occupation", "עיסוק בעל"), ("representative", "שם נציג"),
]

_PRIORITY_LABELS = {4: "קבוע", 3: "עדיפות ראשונה", 2: "עדיפות שנייה"}


def _priority_text(rec: Dict) -> str:
    pr = rec.get("priority")
    if pr in _PRIORITY_LABELS:
        return _PRIORITY_LABELS[pr]
    return "חובת בירור" if "בירור" in (rec.get("priority_raw") or "") else ""


def _fmt_date(v) -> str:
    s = str(v or "")
    if len(s) >= 10 and s[4] == "-":
        return f"{s[8:10]}/{s[5:7]}/{s[:4]}"
    return s


def export_full_distribution_to_excel(recipients: List[Dict], dist_date: str,
                                      dist_name: str = "") -> str:
    """Detailed export of a distribution: EVERY recipient field stored in the app,
    plus a 'קיבל חלוקה' column marking that the people listed received. Returns
    the saved file path."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "חלוקה מלאה"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2563EB")
    got_fill = PatternFill("solid", fgColor="16A34A")   # green for 'קיבל חלוקה'
    header_align = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["מס'", "קיבל חלוקה"] + [h for _, h in _FULL_FIELDS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws[1][1].fill = got_fill   # highlight the 'קיבל חלוקה' header
    ws.row_dimensions[1].height = 22

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    normal_fill = PatternFill("solid", fgColor="FFFFFF")
    got_cell_fill = PatternFill("solid", fgColor="DCFCE7")     # green — received
    reserve_cell_fill = PatternFill("solid", fgColor="FEF3C7")  # amber — reserve
    cell_align = Alignment(horizontal="right", vertical="center")
    _DATE_KEYS = {"last_distribution", "next_distribution", "birth_date", "spouse_birth_date"}

    # Alphabetical by name so the exported sheet reads in א-ב order (bug #scgrh).
    # Reserve (standby) people stay grouped at the end, each group sorted by name.
    # Hebrew sorts correctly by Unicode code point.
    recipients = sorted(recipients,
                        key=lambda r: (bool(r.get("_reserve")),
                                       (r.get("full_name") or "").strip()))

    for i, rec in enumerate(recipients, 1):
        is_reserve = bool(rec.get("_reserve"))
        got = "רזרבה" if is_reserve else "כן"
        row = [i, got]
        for key, _ in _FULL_FIELDS:
            if key == "priority":
                row.append(_priority_text(rec))
            elif key in _DATE_KEYS:
                row.append(_fmt_date(rec.get(key)))
            else:
                v = rec.get(key)
                row.append("" if v is None else v)
        ws.append(row)
        fill = alt_fill if i % 2 == 0 else normal_fill
        for cell in ws[i + 1]:
            cell.alignment = cell_align
            cell.fill = fill
            cell.border = border
        # 'קיבל חלוקה' cell — green for received, amber for reserve (standby)
        ws[i + 1][1].fill = reserve_cell_fill if is_reserve else got_cell_fill
        for c, h in enumerate(headers, 1):   # phones as real numbers (no green triangle)
            if h.startswith("טלפון"):
                _write_phone_cell(ws.cell(i + 1, c), rec.get(_FULL_FIELDS[c - 3][0], ""))
        ws.row_dimensions[i + 1].height = 18

    # column widths — index, got, then a sensible width per field
    widths = [6, 12] + [26 if k == "address" else 20 if k in ("full_name", "email", "synagogue")
                        else 14 for k, _ in _FULL_FIELDS]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    ncols = len(headers)
    last_col = ws.cell(1, ncols).column_letter
    ws.insert_rows(1)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    _name_part = f"{dist_name} — " if dist_name else ""
    title_cell.value = f"{_name_part}רשימת חלוקה מלאה — {dist_date}  (המופיעים קיבלו חלוקה)"
    title_cell.font = Font(bold=True, size=14, color="1D4ED8")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"   # keep title + header visible while scrolling

    exports_dir = _downloads_dir()
    _safe = "".join(c for c in (dist_name or "חלוקה_מלאה") if c not in '\\/:*?"<>|').strip() or "חלוקה_מלאה"
    filename = f"{_safe}_{dist_date.replace('/', '-')}_{datetime.now().strftime('%H%M%S')}.xlsx"
    path = str(exports_dir / filename)
    wb.save(path)
    return path


def export_recipients_to_excel(recipients: List[Dict]) -> str:
    """Export a (filtered) recipients list — every stored field — to Downloads.
    Used by the search screen's 'ייצוא לאקסל'. Reuses _FULL_FIELDS. Returns path."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "מקבלים"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["מס'"] + [h for _, h in _FULL_FIELDS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[1].height = 22

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    normal_fill = PatternFill("solid", fgColor="FFFFFF")
    cell_align = Alignment(horizontal="right", vertical="center")
    _DATE_KEYS = {"last_distribution", "next_distribution", "birth_date", "spouse_birth_date"}

    for i, rec in enumerate(recipients, 1):
        row = [i]
        for key, _ in _FULL_FIELDS:
            if key == "priority":
                row.append(_priority_text(rec))
            elif key in _DATE_KEYS:
                row.append(_fmt_date(rec.get(key)))
            else:
                v = rec.get(key)
                row.append("" if v is None else v)
        ws.append(row)
        fill = alt_fill if i % 2 == 0 else normal_fill
        for cell in ws[i + 1]:
            cell.alignment = cell_align
            cell.fill = fill
            cell.border = border
        ws.row_dimensions[i + 1].height = 18

    widths = [6] + [26 if k == "address" else 20 if k in ("full_name", "email", "synagogue")
                    else 14 for k, _ in _FULL_FIELDS]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ncols = len(headers)
    ws.insert_rows(1)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    title_cell = ws["A1"]
    title_cell.value = f"רשימת מקבלים — {datetime.now().strftime('%d/%m/%Y')}  ({len(recipients)})"
    title_cell.font = Font(bold=True, size=14, color="1D4ED8")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"

    exports_dir = _downloads_dir()
    filename = f"מקבלים_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    path = str(exports_dir / filename)
    wb.save(path)
    return path


# ─── Volunteer checklist (send-out / read-back round trip) ────────────────────
# A minimal, privacy-conscious checklist a volunteer fills WITHOUT touching the
# app: mark who came, a note per recipient, and one general note for the whole
# round. A hidden hidden "meta" sheet carries the round's date/what/qty/
# distributor/dist_name and the recipient id per row, so re-import needs no
# re-typing and matches rows reliably even if the volunteer reorders/deletes
# some rows.

_VOL_HEADER_ROW = 9
_VOL_DATA_START_ROW = 10
_VOL_COLS = ["מזהה", "מס'", "שם מלא", "טלפון 1", "טלפון 2", "טלפון 3",
            "אזור", "נפשות", "הגיע?", "הערה למקבל"]
_VOL_COL_ID, _VOL_COL_NUM, _VOL_COL_NAME = 1, 2, 3
_VOL_COL_PHONE1, _VOL_COL_PHONE2, _VOL_COL_PHONE3 = 4, 5, 6
_VOL_COL_AREA, _VOL_COL_SOULS, _VOL_COL_CAME, _VOL_COL_NOTE = 7, 8, 9, 10
_VOL_GENERAL_NOTE_ROW = 6   # top-left of the merged general-note input block


def export_volunteer_checklist_to_excel(recipients: List[Dict], dist_date: str,
                                        dist_name: str, what: str, qty,
                                        distributor: str) -> str:
    """Build a nicely styled, minimal checklist for a volunteer to fill by hand
    (no app access needed): who came, a note per person, and one general note
    for the round. Only the fields a volunteer needs are included — NOT the
    full sensitive recipient record. Returns the saved file path."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "רשימה למתנדב"
    ws.sheet_view.rightToLeft = True

    blue = "1565C0"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=blue)
    header_align = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Branded header block (text-based — no image dependency) ────────────
    last_col_letter = get_column_letter(len(_VOL_COLS))
    ws.merge_cells(f"A1:{last_col_letter}1")
    org_cell = ws["A1"]
    org_cell.value = "קופה של צדקה הר יונה — נוף הגליל"
    org_cell.font = Font(bold=True, size=13, color=blue)
    org_cell.alignment = Alignment(horizontal="center", vertical="center")
    org_cell.fill = PatternFill("solid", fgColor="EEF4FF")
    ws.row_dimensions[1].height = 26

    dist_date_disp = _fmt_date(dist_date)

    ws.merge_cells(f"A2:{last_col_letter}2")
    title_cell = ws["A2"]
    title_cell.value = f"{dist_name} — {dist_date_disp}"
    title_cell.font = Font(bold=True, size=16, color="0D2A4A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    ws.merge_cells(f"A3:{last_col_letter}3")
    sub_cell = ws["A3"]
    qty_txt = f"  ·  כמות: {qty}" if qty else ""
    sub_cell.value = f"מה חולק: {what}{qty_txt}  ·  מחלק: {distributor}"
    sub_cell.font = Font(size=11, color="475569")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    instr = ws["A4"]
    ws.merge_cells(f"A4:{last_col_letter}4")
    instr.value = ("הוראות: כל השמות מסומנים מראש \"כן\" (קיבל). יש לשנות ל\"לא\" רק את מי שלא הגיע. "
                   "אפשר להוסיף הערה פרטנית לכל אחד. בסוף — למלא הערה כללית למטה ולשלוח את הקובץ חזרה.")
    instr.font = Font(size=10, italic=True, color="6B7280")
    instr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    # ── General note block (row 5 label + rows 6-7 merged input) ───────────
    ws.merge_cells(f"B5:{last_col_letter}5")
    note_label = ws["B5"]
    note_label.value = "הערה כללית על כל החלוקה (למלא בסוף):"
    note_label.font = Font(bold=True, size=11, color="92400E")
    note_label.alignment = Alignment(horizontal="right", vertical="center")
    note_label.fill = PatternFill("solid", fgColor="FEF3C7")

    ws.merge_cells(f"B{_VOL_GENERAL_NOTE_ROW}:{last_col_letter}{_VOL_GENERAL_NOTE_ROW + 1}")
    note_input = ws.cell(_VOL_GENERAL_NOTE_ROW, 2)
    note_input.fill = PatternFill("solid", fgColor="FFFBEB")
    note_input.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    note_input.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[_VOL_GENERAL_NOTE_ROW].height = 22
    ws.row_dimensions[_VOL_GENERAL_NOTE_ROW + 1].height = 22

    ws.row_dimensions[8].height = 8   # spacer

    # ── Checklist table ──────────────────────────────────────────────────────
    for c, h in enumerate(_VOL_COLS, 1):
        cell = ws.cell(_VOL_HEADER_ROW, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[_VOL_HEADER_ROW].height = 22

    came_options = '"כן,לא"'
    dv = DataValidation(type="list", formula1=came_options, allow_blank=True,
                        showDropDown=False)
    ws.add_data_validation(dv)

    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    normal_fill = PatternFill("solid", fgColor="FFFFFF")
    cell_align = Alignment(horizontal="right", vertical="center")

    for i, rec in enumerate(recipients, 1):
        r = _VOL_DATA_START_ROW + i - 1
        # "הגיע?" defaults to "כן" — the volunteer only changes the few who didn't.
        vals = [rec.get("id", ""), i, rec.get("full_name", ""),
                rec.get("phone1", ""), rec.get("phone2", ""), rec.get("phone3", ""),
                rec.get("area", ""), rec.get("souls", ""), "כן", ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = cell_align
            cell.border = border
            cell.fill = alt_fill if i % 2 == 0 else normal_fill
        for c, h in enumerate(_VOL_COLS, 1):   # phones as real numbers (no green triangle)
            if h.startswith("טלפון"):
                _write_phone_cell(ws.cell(r, c), vals[c - 1])
        ws.cell(r, _VOL_COL_NAME).font = Font(bold=True)
        dv.add(ws.cell(r, _VOL_COL_CAME))
        ws.row_dimensions[r].height = 18

    widths = {1: 8, 2: 6, 3: 24, 4: 15, 5: 15, 6: 15, 7: 10, 8: 8, 9: 10, 10: 28}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.column_dimensions[get_column_letter(_VOL_COL_ID)].hidden = True
    ws.freeze_panes = ws.cell(_VOL_HEADER_ROW + 1, 1).coordinate

    # ── Hidden "meta" sheet — round details, read back automatically on import
    meta = wb.create_sheet("meta")
    meta["A1"], meta["B1"] = "dist_date", dist_date
    meta["A2"], meta["B2"] = "what", what
    meta["A3"], meta["B3"] = "qty", qty
    meta["A4"], meta["B4"] = "distributor", distributor
    meta["A5"], meta["B5"] = "dist_name", dist_name
    meta.sheet_state = "hidden"

    exports_dir = _downloads_dir()
    _safe = "".join(c for c in (dist_name or "רשימה_למתנדב") if c not in '\\/:*?"<>|').strip() or "רשימה_למתנדב"
    filename = f"{_safe}_{dist_date.replace('/', '-')}_{datetime.now().strftime('%H%M%S')}.xlsx"
    path = str(exports_dir / filename)
    wb.save(path)
    # Lock the file with an open-password if one is configured (the volunteer
    # gets the password in the email body). No-op when no password is set.
    _encrypt_xlsx(path, _checklist_password())
    return path


def import_volunteer_checklist(path: str) -> dict:
    """Read a volunteer's filled-in checklist (as produced by
    export_volunteer_checklist_to_excel). Returns:
    {
      "meta": {dist_date, what, qty, distributor, dist_name},
      "received": [ {id, full_name, area, souls, frequency, notes}, ... ]  # marked כן
      "unmatched": [ row-name strings that could not be matched to a recipient ]
    }
    Matching: primarily by the hidden recipient id column; if that id doesn't
    exist in the DB (edited/blank), falls back to an exact full_name match.
    """
    import database as db

    wb = _load_maybe_encrypted(path, data_only=True)
    ws = wb.worksheets[0]

    meta = {"dist_date": "", "what": "", "qty": 0, "distributor": "", "dist_name": "",
            "general_note": ""}
    if "meta" in wb.sheetnames:
        m = wb["meta"]
        for row in m.iter_rows(min_row=1, max_row=5, values_only=True):
            if not row or row[0] is None:
                continue
            key, val = row[0], row[1]
            if key in meta:
                meta[key] = val if val is not None else ""
    try:
        meta["qty"] = int(meta["qty"]) if meta["qty"] not in ("", None) else 0
    except (TypeError, ValueError):
        meta["qty"] = 0

    general_note = ws.cell(_VOL_GENERAL_NOTE_ROW, 2).value
    meta["general_note"] = (str(general_note).strip() if general_note else "")

    received = []
    unmatched = []
    for r in range(_VOL_DATA_START_ROW, ws.max_row + 1):
        name_cell = ws.cell(r, _VOL_COL_NAME).value
        id_cell = ws.cell(r, _VOL_COL_ID).value
        if name_cell is None and id_cell is None:
            continue   # a row the volunteer emptied/deleted — just skip it

        # "הגיע?" defaults to "כן"; a recipient counts as RECEIVED unless the
        # volunteer explicitly marked "לא" (or cleared/crossed out the cell).
        came = str(ws.cell(r, _VOL_COL_CAME).value or "").strip()
        did_not_come = came in ("", "לא", "X", "x", "✗", "-", "no", "No")
        if not did_not_come:
            rec = None
            try:
                rid = int(id_cell) if id_cell not in (None, "") else None
            except (TypeError, ValueError):
                rid = None
            if rid is not None:
                rec = db.get_recipient(rid)
            if rec is None and name_cell:
                candidates = [x for x in db.get_all_recipients()
                             if (x.get("full_name") or "").strip() == str(name_cell).strip()]
                rec = candidates[0] if len(candidates) == 1 else None
            if rec is None:
                unmatched.append(str(name_cell or f"שורה {r}"))
            else:
                note_cell = ws.cell(r, _VOL_COL_NOTE).value
                received.append({
                    "id": rec["id"], "full_name": rec.get("full_name", ""),
                    "area": rec.get("area", ""), "souls": rec.get("souls", 0),
                    "frequency": rec.get("frequency", ""),
                    "notes": str(note_cell).strip() if note_cell else "",
                })

    return {"meta": meta, "received": received, "unmatched": unmatched}
